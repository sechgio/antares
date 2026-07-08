"""Genera la plantilla Excel del Sheet maestro AUTOIMG (6 pestañas obligatorias)."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "formatos" / "AUTOIMG_plantilla_maestro.xlsx"

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
BODY_FONT = Font(name="Arial", size=10)
NOTE_FONT = Font(name="Arial", size=10, italic=True, color="666666")
WRAP = Alignment(wrap_text=True, vertical="top")


def _style_header(ws: Worksheet, headers: list[str], widths: list[int] | None = None) -> None:
    ws.append(headers)
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        width = widths[col_idx - 1] if widths and col_idx - 1 < len(widths) else max(12, len(title) + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


def _append_rows(ws: Worksheet, rows: list[list[object]]) -> None:
    for row in rows:
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.font = BODY_FONT
            cell.alignment = WRAP


def _build_instructions(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("INSTRUCCIONES", 0)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 72
    lines = [
        ("Plantilla AUTOIMG — Sheet maestro", ""),
        ("", ""),
        ("Uso", "Sube este archivo a Google Drive y ábrelo como Google Sheet, o crea un Sheet vacío y copia cada pestaña con los mismos nombres y columnas."),
        ("Vincular en ANTARES", "AutoIMG → panel lateral → Google Sheets → pegar URL/ID del Sheet → Abrir. La app crea pestañas faltantes automáticamente."),
        ("", ""),
        ("Pestañas obligatorias (6)", "BD_IMG, FOLDERS, BD_ARRASTRE, LOGS, CONFIG, RESUMEN"),
        ("BD_IMG", "Padrón de NIS (7 dígitos en nombres de imagen). El escaneo actualiza IMG_1..3, CANTIDAD, ESTADO y ORIGEN_CARPETAS."),
        ("FOLDERS", "Carpetas de Google Drive a escanear. ACTIVO: ✅ = escanea, ❌ = ignora. También puedes agregar carpetas desde la app."),
        ("BD_ARRASTRE", "Casos manuales de arrastre/reasignación. Solo lectura en la app."),
        ("LOGS", "Historial de operaciones (la app escribe SCAN_ALL_FOLDERS tras cada sincronización)."),
        ("CONFIG", "Claves: DEDUP_STRATEGY (SUM|MAX), AUTO_SYNC (true|false), ULTIMO_SYNC, SHEET_ID, USUARIO."),
        ("RESUMEN", "Métricas del padrón. La app recalcula tras cada escaneo+sincronización."),
        ("", ""),
        ("Estados de imagen", "🟢 COMPLETO = 3 imágenes · 🔴 FALTANTE = menos de 3 · 🟡 SOBRANTE = más de 3"),
        ("NIS en archivos", "El nombre debe contener un número de 7 dígitos, ej.: 4210801-foto.jpg"),
        ("", ""),
        ("Filas de ejemplo", "Las pestañas de datos incluyen filas de muestra. Bórralas antes de producción o déjalas como referencia."),
    ]
    for row_idx, (left, right) in enumerate(lines, start=1):
        ws.cell(row=row_idx, column=1, value=left).font = Font(name="Arial", bold=bool(left and not right), size=11 if row_idx == 1 else 10)
        cell_b = ws.cell(row=row_idx, column=2, value=right)
        cell_b.font = NOTE_FONT if row_idx > 1 else BODY_FONT
        cell_b.alignment = WRAP


def _build_bd_img(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("BD_IMG")
    headers = [
        "NIS", "SGIO", "DESTINO", "NOMBRE", "DIRECCION",
        "IMG_1", "IMG_2", "IMG_3", "CANTIDAD", "ESTADO",
        "ORIGEN_CARPETAS", "ULTIMA_VERIFICACION", "NOTAS",
    ]
    widths = [12, 12, 14, 22, 28, 8, 8, 8, 10, 16, 24, 20, 24]
    _style_header(ws, headers, widths)
    _append_rows(ws, [
        [
            "4210801", "69656525", "DVD 03", "Cliente completo", "Av. Principal 123",
            "✅", "✅", "✅", 3, "🟢 COMPLETO", "CARPETA_EJEMPLO", "", "",
        ],
        [
            "4210802", "", "DVD 03", "Cliente faltante", "Jr. Los Olivos 456",
            "✅", "⬜", "⬜", 1, "🔴 FALTANTE", "CARPETA_EJEMPLO", "", "NUEVO (sin SGIO)",
        ],
        [
            "4210803", "69656527", "DVD 03", "Cliente sobrante", "Calle Secundaria 789",
            "✅", "✅", "✅", 4, "🟡 SOBRANTE", "CARPETA_A; CARPETA_B", "", "",
        ],
    ])


def _build_folders(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("FOLDERS")
    headers = ["NOMBRE", "FOLDER_ID", "ACTIVO", "ULTIMO_SCAN", "CANT_ARCHIVOS"]
    widths = [24, 36, 10, 20, 14]
    _style_header(ws, headers, widths)
    _append_rows(ws, [
        ["CARPETA_EJEMPLO", "REEMPLAZAR_CON_FOLDER_ID_DE_DRIVE", "✅", "", 0],
        ["CARPETA_INACTIVA", "REEMPLAZAR_OPCIONAL", "❌", "", 0],
    ])


def _build_arrastre(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("BD_ARRASTRE")
    headers = ["NIS", "SGIO", "MOTIVO", "FECHA", "OBSERVACION"]
    widths = [12, 12, 24, 14, 36]
    _style_header(ws, headers, widths)
    _append_rows(ws, [
        ["4210801", "69656525", "Arrastre manual", "2026-07-01", "Ejemplo: reasignación de expediente"],
    ])


def _build_logs(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("LOGS")
    headers = ["FECHA", "ACCION", "DETALLE", "USUARIO", "DURACION"]
    widths = [20, 18, 48, 28, 10]
    _style_header(ws, headers, widths)
    _append_rows(ws, [
        [
            "2026-07-06 10:00:00",
            "SCAN_ALL_FOLDERS",
            "1 carpetas · 3 NIS · 1 actualizados · 2 nuevos",
            "usuario@ejemplo.com",
            "12.5",
        ],
    ])


def _build_config(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("CONFIG")
    headers = ["Clave", "Valor"]
    widths = [22, 48]
    _style_header(ws, headers, widths)
    _append_rows(ws, [
        ["DEDUP_STRATEGY", "SUM"],
        ["AUTO_SYNC", "false"],
        ["ULTIMO_SYNC", ""],
        ["SHEET_ID", ""],
        ["USUARIO", ""],
    ])
    note_row = ws.max_row + 2
    ws.cell(row=note_row, column=1, value="Notas").font = Font(name="Arial", bold=True, size=10)
    ws.cell(row=note_row + 1, column=1, value="DEDUP_STRATEGY").font = BODY_FONT
    ws.cell(row=note_row + 1, column=2, value="SUM = suma imágenes entre carpetas (default). MAX = toma el máximo por NIS.").font = NOTE_FONT
    ws.cell(row=note_row + 2, column=1, value="AUTO_SYNC").font = BODY_FONT
    ws.cell(row=note_row + 2, column=2, value="true = la app recarga el Sheet cada 5 min (no re-escanea Drive).").font = NOTE_FONT


def _build_resumen(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("RESUMEN")
    headers = ["METRICA", "VALOR", "FECHA"]
    widths = [28, 14, 20]
    _style_header(ws, headers, widths)
    _append_rows(ws, [
        ["TOTAL NIS", "3", ""],
        ["🟢 COMPLETOS (3/3)", "1", ""],
        ["🔴 FALTANTES (<3)", "1", ""],
        ["🟡 SOBRANTES (>3)", "1", ""],
        ["SIN SGIO", "1", ""],
        ["CARPETAS ACTIVAS", "1", ""],
        ["ULTIMO PROCESO", "", ""],
    ])


def main() -> None:
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    _build_instructions(wb)
    _build_bd_img(wb)
    _build_folders(wb)
    _build_arrastre(wb)
    _build_logs(wb)
    _build_config(wb)
    _build_resumen(wb)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"[OK] Plantilla generada: {OUTPUT}")


if __name__ == "__main__":
    main()
