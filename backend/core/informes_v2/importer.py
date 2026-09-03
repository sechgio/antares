from __future__ import annotations

import csv
import io
import re
import unicodedata
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook

from backend.core.informes_v2.models import (
    DIAMETERS,
    LINEA_ROWS,
    VALVULA_ROWS,
    InformeV2,
    _safe_int,
    _safe_str,
    create_empty_report,
    report_id_from_number,
)

COLUMN_MAPPING: dict[str, str] = {
    "nroinforme": "informe_id",
    "numeroinforme": "informe_id",
    "informeid": "informe_id",
    "informe": "informe_id",
    "item": "informe_id",
    "nro": "informe_id",
    "numero": "informe_id",
    "id": "photo_id",
    "codigoid": "photo_id",
    "photoid": "photo_id",
    "codigo": "photo_id",
    "codigoimagen": "photo_id",
    "estacion": "estacion",
    "estacionnombre": "estacion",
    "nombreestacion": "estacion",
    "reservorio": "estacion",
    "tipo": "tipo",
    "tipoestructura": "tipo",
    "tiporeservorio": "tipo",
    "volumen": "volumen",
    "volumenm3": "volumen",
    "capacidad": "volumen",
    "ubicacion": "ubicacion",
    "direccion": "ubicacion",
    "lugar": "ubicacion",
    "distrito": "distrito",
    "fecha": "fecha_ejecucion",
    "fechaejecucion": "fecha_ejecucion",
    "fechadeejecucion": "fecha_ejecucion",
    "suministro": "suministro",
    "nrosuministro": "suministro",
    "nis": "suministro",
    "sgio": "sgio",
    "nrosgio": "sgio",
    "observacion": "medidas_observacion",
    "observaciones": "medidas_observacion",
    "obs": "medidas_observacion",
}

VALVULA_ALIASES = {
    "conduccion": ("conduccion", "cond", "valvcond", "valvulasconduccion"),
    "impulsion": ("impulsion", "imp", "valvimp", "valvulasimpulsion"),
    "aduccion": ("aduccion", "aduc", "valvaduc", "valvulasaduccion"),
    "bypass": ("bypass", "pass", "by", "valvbypass", "valvulasbypass"),
    "purga": ("purga", "valvpurga", "valvulaspurga"),
}

LINEA_ALIASES = {
    "aduccion": ("lineaaduccion", "tuberiaaduccion", "aduccionlinea"),
    "alimentacion": ("alimentacion", "lineaalimentacion", "aliment"),
    "impulsion_rebombeo": (
        "impulsionrebombeo",
        "impulsionebombeo",
        "rebombeo",
        "lineaimpulsion",
        "linimpulsion",
        "impulsion",
    ),
    "rebose": ("rebose", "linearebose"),
    "purga": ("lineapurga", "purgalinea"),
}

for section, aliases in VALVULA_ALIASES.items():
    for alias in aliases:
        for diameter in DIAMETERS:
            COLUMN_MAPPING[f"{alias}{diameter}"] = f"valvulas_{section}_{diameter}"
            COLUMN_MAPPING[f"valv{alias}{diameter}"] = f"valvulas_{section}_{diameter}"
            COLUMN_MAPPING[f"valvulas{alias}{diameter}"] = f"valvulas_{section}_{diameter}"
        COLUMN_MAPPING[f"{alias}oper"] = f"valvulas_{section}_oper"
        COLUMN_MAPPING[f"{alias}noop"] = f"valvulas_{section}_no_op"
        COLUMN_MAPPING[f"{alias}nooper"] = f"valvulas_{section}_no_op"
        COLUMN_MAPPING[f"{alias}obs"] = f"valvulas_{section}_obs"
        COLUMN_MAPPING[f"obs{alias}"] = f"valvulas_{section}_obs"
        COLUMN_MAPPING[f"observaciones{alias}"] = f"valvulas_{section}_obs"

for section, aliases in LINEA_ALIASES.items():
    for alias in aliases:
        for diameter in DIAMETERS:
            COLUMN_MAPPING[f"{alias}{diameter}"] = f"linea_{section}_{diameter}"
            COLUMN_MAPPING[f"linea{alias}{diameter}"] = f"linea_{section}_{diameter}"
        COLUMN_MAPPING[f"{alias}oper"] = f"linea_{section}_oper"
        COLUMN_MAPPING[f"{alias}noop"] = f"linea_{section}_no_op"
        COLUMN_MAPPING[f"{alias}nooper"] = f"linea_{section}_no_op"
        COLUMN_MAPPING[f"{alias}obs"] = f"linea_{section}_obs"
        COLUMN_MAPPING[f"obs{alias}"] = f"linea_{section}_obs"

for section in VALVULA_ROWS:
    for diameter in DIAMETERS:
        COLUMN_MAPPING[f"valv{section}{diameter}"] = f"valvulas_{section}_{diameter}"
    COLUMN_MAPPING[f"valv{section}oper"] = f"valvulas_{section}_oper"
    COLUMN_MAPPING[f"valv{section}noop"] = f"valvulas_{section}_no_op"
    COLUMN_MAPPING[f"valv{section}obs"] = f"valvulas_{section}_obs"

for section in LINEA_ROWS:
    short = section.replace("_", "")
    for diameter in DIAMETERS:
        COLUMN_MAPPING[f"lin{short}{diameter}"] = f"linea_{section}_{diameter}"
        COLUMN_MAPPING[f"linea{short}{diameter}"] = f"linea_{section}_{diameter}"
    COLUMN_MAPPING[f"lin{short}oper"] = f"linea_{section}_oper"
    COLUMN_MAPPING[f"lin{short}noop"] = f"linea_{section}_no_op"
    COLUMN_MAPPING[f"lin{short}obs"] = f"linea_{section}_obs"

MEDIDA_ALIASES = {
    "largo": ("largo", "medidalargo", "longitud"),
    "ancho": ("ancho", "medidaancho", "width"),
    "diametro": ("diametro", "medidadiametro"),
    "altura_rebose": ("alturarebose", "medidaalturarebose", "rebosealtura"),
    "altura_total": ("alturatotal", "medidaalturatotal"),
    "tirante_limpieza": ("tirantelimpeza", "tirantelimpieza", "medidatirante"),
}

for field, aliases in MEDIDA_ALIASES.items():
    for alias in aliases:
        COLUMN_MAPPING[alias] = f"medidas_{field}"


def normalize_header_value(value: str) -> str:
    if not value:
        return ""
    text = str(value).strip().lower().replace("\ufeff", "")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"\b(de|del|la|el)\b", "", text)
    return re.sub(r"[\s_\.:\-°/()\"']+", "", text)


def normalize_csv_key(value: str) -> str:
    mapped = COLUMN_MAPPING.get(normalize_header_value(value))
    if mapped:
        return mapped
    text = str(value or "").strip().lower().replace("\ufeff", "")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def import_reports_from_bytes(filename: str, content: bytes) -> list[dict[str, Any]]:
    lower_name = filename.lower()
    if lower_name.endswith(".csv"):
        rows = parse_csv_file(content)
    elif lower_name.endswith(".xlsx") or lower_name.endswith(".xls"):
        rows = parse_xlsx_file(content)
    else:
        msg = "Formato no soportado. Use archivos .csv o .xlsx"
        raise ValueError(msg)
    if not rows:
        msg = "El archivo esta vacio o no tiene datos validos"
        raise ValueError(msg)

    reports: list[dict[str, Any]] = []
    used_numbers: set[int] = set()
    next_report_number = 1

    for row in rows:
        explicit_number = _safe_int(row.get("informe_id"), 0)
        if explicit_number > 0 and explicit_number not in used_numbers:
            report_number = explicit_number
        else:
            while next_report_number in used_numbers:
                next_report_number += 1
            report_number = next_report_number

        used_numbers.add(report_number)
        next_report_number = max(next_report_number, report_number) + 1
        reports.append(transform_flat_to_nested(row, report_number))

    return reports


def parse_csv_file(content: bytes) -> list[dict[str, Any]]:
    text: str | None = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252", "iso-8859-1"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue

    if text is None:
        msg = "No se pudo decodificar el archivo CSV"
        raise ValueError(msg)

    def _read_rows(delimiter: str) -> list[dict[str, Any]]:
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        parsed_rows: list[dict[str, Any]] = []
        for row in reader:
            normalized = {normalize_csv_key(key): value for key, value in row.items() if key is not None}
            if any(str(value or "").strip() for value in normalized.values()):
                parsed_rows.append(normalized)
        return parsed_rows

    semicolon_rows = _read_rows(";")
    comma_rows = _read_rows(",")
    tab_rows = _read_rows("\t")

    def _mapped_col_count(rows: list[dict[str, Any]]) -> int:
        if not rows:
            return 0
        return len([k for k in rows[0] if k and k in COLUMN_MAPPING.values()])

    semi_cols = _mapped_col_count(semicolon_rows)
    comma_cols = _mapped_col_count(comma_rows)
    tab_cols = _mapped_col_count(tab_rows)
    best_cols = max(semi_cols, comma_cols, tab_cols)
    if best_cols >= 1:
        if semi_cols == best_cols and semicolon_rows:
            return semicolon_rows
        if tab_cols == best_cols and tab_rows:
            return tab_rows
        if comma_rows:
            return comma_rows
    return semicolon_rows or comma_rows or tab_rows


def parse_xlsx_file(content: bytes) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as err:
        msg = f"Error al abrir el archivo Excel: {err}"
        raise ValueError(msg) from err

    sheets = workbook.worksheets
    if not sheets:
        return []

    best_rows: list[dict[str, Any]] = []
    best_sheet_score = -1

    for worksheet in sheets:
        rows_iter = list(worksheet.iter_rows(values_only=True))
        if not rows_iter:
            continue

        header_idx = -1
        best_score = -1
        for idx, row_tuple in enumerate(rows_iter[:10]):
            non_empty = [v for v in row_tuple if v is not None and str(v).strip() != ""]
            if not non_empty:
                continue
            score = sum(1 for v in non_empty if normalize_header_value(str(v)) in COLUMN_MAPPING)
            if score > best_score:
                best_score = score
                header_idx = idx

        if header_idx == -1 or best_score <= 0:
            for idx, row_tuple in enumerate(rows_iter[:10]):
                if any(v is not None and str(v).strip() != "" for v in row_tuple):
                    header_idx = idx
                    break

        if header_idx == -1:
            continue

        headers = rows_iter[header_idx]
        keys = [normalize_csv_key(str(header or "")) for header in headers]
        rows: list[dict[str, Any]] = []
        consecutive_empty = 0
        for values in rows_iter[header_idx + 1 :]:
            row_dict: dict[str, Any] = {
                keys[idx]: values[idx] for idx in range(min(len(keys), len(values))) if keys[idx]
            }
            if any(value is not None and str(value or "").strip() != "" for value in row_dict.values()):
                rows.append(row_dict)
                consecutive_empty = 0
            else:
                consecutive_empty += 1
                if consecutive_empty >= 50:
                    break

        if rows and best_score > best_sheet_score:
            best_sheet_score = best_score
            best_rows = rows

    return best_rows


def transform_flat_to_nested(row: dict[str, Any], fallback_report_number: int = 1) -> dict[str, Any]:
    informe_id = _safe_int(row.get("informe_id"), 0)
    if informe_id <= 0:
        informe_id = fallback_report_number

    report = create_empty_report(informe_id)
    report["id"] = report_id_from_number(informe_id)

    photo_id = _safe_str(row.get("photo_id"))
    estacion = _safe_str(row.get("estacion"))
    suministro = _safe_str(row.get("suministro"))
    report["header"].update(
        {
            "photo_id": photo_id or estacion or suministro or report["id"],
            "estacion": estacion,
            "tipo": _safe_str(row.get("tipo"), "ELEVADO").upper(),
            "volumen": _safe_int(row.get("volumen"), 0),
            "ubicacion": _safe_str(row.get("ubicacion")),
            "distrito": _safe_str(row.get("distrito")),
            "fecha_ejecucion": _format_fecha(row.get("fecha_ejecucion")),
            "suministro": suministro,
            "sgio": _safe_str(row.get("sgio")),
        }
    )

    for section in VALVULA_ROWS:
        diametros = {d: _safe_int(row.get(f"valvulas_{section}_{d}"), 0) for d in DIAMETERS}
        report["valvulas"][section] = {
            "diametros": diametros,
            "oper": _safe_int(row.get(f"valvulas_{section}_oper"), 0),
            "no_op": _safe_int(row.get(f"valvulas_{section}_no_op"), 0),
            "observaciones": _safe_str(row.get(f"valvulas_{section}_obs")),
        }

    for section in LINEA_ROWS:
        diametros = {d: _safe_int(row.get(f"linea_{section}_{d}"), 0) for d in DIAMETERS}
        report["linea"][section] = {
            "diametros": diametros,
            "oper": _safe_int(row.get(f"linea_{section}_oper"), 0),
            "no_op": _safe_int(row.get(f"linea_{section}_no_op"), 0),
            "observaciones": _safe_str(row.get(f"linea_{section}_obs")),
        }

    report["medidas"].update(
        {
            "largo": _safe_str(row.get("medidas_largo")),
            "ancho": _safe_str(row.get("medidas_ancho")),
            "diametro": _safe_str(row.get("medidas_diametro")),
            "altura_rebose": _safe_str(row.get("medidas_altura_rebose")),
            "altura_total": _safe_str(row.get("medidas_altura_total")),
            "tirante_limpieza": _safe_str(row.get("medidas_tirante_limpieza")),
            "observacion": _safe_str(row.get("medidas_observacion")),
        }
    )
    report["last_modified"] = datetime.now().isoformat()
    return InformeV2.normalize(report)


def _format_fecha(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%d/%m/%Y")
    text = str(value).strip()
    m = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})", text)
    if m:
        y, month, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{d:02d}/{month:02d}/{y}"
    return text


TEMPLATE_HEADERS: list[str] = [
    "ID",
    "Informe",
    "Estacion",
    "Tipo",
    "Volumen",
    "Ubicacion",
    "Distrito",
    "Fecha Ejecucion",
    "Suministro",
    "SGIO",
]

for _section, label in [
    ("conduccion", "Valv Cond"),
    ("impulsion", "Valv Imp"),
    ("aduccion", "Valv Aduc"),
    ("bypass", "Valv Bypass"),
    ("purga", "Valv Purga"),
]:
    for diameter in DIAMETERS:
        TEMPLATE_HEADERS.append(f"{label} {diameter}")
    TEMPLATE_HEADERS.append(f"{label} Oper")
    TEMPLATE_HEADERS.append(f"{label} No Op")
    TEMPLATE_HEADERS.append(f"{label} Obs")

for _section, label in [
    ("aduccion", "Lin Aduccion"),
    ("alimentacion", "Lin Alimentacion"),
    ("impulsion_rebombeo", "Lin Impulsion"),
    ("rebose", "Lin Rebose"),
    ("purga", "Lin Purga"),
]:
    for diameter in DIAMETERS:
        TEMPLATE_HEADERS.append(f"{label} {diameter}")
    TEMPLATE_HEADERS.append(f"{label} Oper")
    TEMPLATE_HEADERS.append(f"{label} No Op")
    TEMPLATE_HEADERS.append(f"{label} Obs")

TEMPLATE_HEADERS.extend(
    [
        "Largo",
        "Ancho",
        "Diametro",
        "Altura Rebose",
        "Altura Total",
        "Tirante Limpieza",
        "Observacion",
    ]
)
