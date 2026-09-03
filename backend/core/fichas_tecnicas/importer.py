
from __future__ import annotations

import csv
import io
import unicodedata
from datetime import datetime
from typing import Any

from openpyxl import load_workbook

from backend.core.fichas_tecnicas.models import (
    FichaTecnica,
    _normalize_satisfaccion,
    _safe_bool,
    _safe_str,
    create_empty_ficha,
    ficha_id_from_number,
)


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value)).encode("ASCII", "ignore").decode("utf-8")
    return text.strip().lower().replace("\ufeff", "").replace(" ", "_")


def parse_csv_bytes(content: bytes) -> list[dict[str, Any]]:
    decoded: str | None = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            decoded = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        msg = "No se pudo decodificar el archivo CSV"
        raise ValueError(msg)

    def _read(delimiter: str) -> list[dict[str, Any]]:
        return list(csv.DictReader(io.StringIO(decoded), delimiter=delimiter))

    semicolon_rows = _read(";")
    comma_rows = _read(",")
    semi_cols = len([k for k in (semicolon_rows[0].keys() if semicolon_rows else []) if k])
    comma_cols = len([k for k in (comma_rows[0].keys() if comma_rows else []) if k])
    rows = semicolon_rows if semi_cols >= comma_cols and semi_cols >= 2 else comma_rows

    cleaned: list[dict[str, Any]] = []
    for row in rows:
        cleaned_row: dict[str, Any] = {}
        has_content = False
        for key, value in row.items():
            if not key:
                continue
            clean_key = _normalize_header(key)
            cleaned_row[clean_key] = value
            if value is not None and str(value).strip() != "":
                has_content = True
        if has_content:
            cleaned.append(cleaned_row)
    return cleaned


def parse_xlsx_bytes(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(content), data_only=True)
    sheet = workbook.active
    if sheet is None:
        return []
    all_rows = list(sheet.iter_rows(values_only=False))
    if not all_rows:
        return []

    headers = [_normalize_header(cell.value) if cell else f"_col_{idx}" for idx, cell in enumerate(all_rows[0])]
    parsed: list[dict[str, Any]] = []
    for row in all_rows[1:]:
        if not any(c is not None and c.value is not None for c in row):
            continue
        row_dict: dict[str, Any] = {}
        has_useful = False
        for col_idx, cell in enumerate(row):
            if col_idx >= len(headers):
                continue
            key = headers[col_idx]
            cell_value = cell.value if cell else None
            if cell_value is None:
                continue
            if "concentracion" in key and isinstance(cell_value, (int, float)) and cell is not None:
                try:
                    fmt = cell.number_format or ""
                    if "0.000" in fmt:
                        cell_value = f"{cell_value:.3f}"
                    elif "0.00" in fmt:
                        cell_value = f"{cell_value:.2f}"
                    elif "0.0" in fmt:
                        cell_value = f"{cell_value:.1f}"
                    elif fmt in {"0", "#"}:
                        cell_value = f"{cell_value:.0f}"
                except (ValueError, TypeError):
                    pass
            row_dict[key] = cell_value
            if str(cell_value).strip():
                has_useful = True
        if has_useful:
            parsed.append(row_dict)
    return parsed


def _row_to_ficha(row: dict[str, Any], number: int) -> dict[str, Any]:
    ficha = create_empty_ficha(number)
    ficha.update(
        {
            "id": ficha_id_from_number(number),
            "os_numero": _safe_str(row.get("os_numero"), f"N° {number:05d}"),
            "cliente": _safe_str(row.get("cliente")),
            "fecha": _safe_str(row.get("fecha")),
            "direccion": _safe_str(row.get("direccion")),
            "distrito": _safe_str(row.get("distrito")),
            "servicio": {
                "desinfeccion": _safe_bool(row.get("servicio_desinfeccion")),
                "limpieza_ambientes": _safe_bool(row.get("servicio_limpieza_ambientes")),
                "limpieza_pozos_septicos": _safe_bool(row.get("servicio_limpieza_pozos")),
                "limpieza_reservorios": _safe_bool(row.get("servicio_limpieza_reservorios")),
            },
            "diagnostico_area": _safe_str(row.get("diagnostico_area")),
            "condicion_sanitaria": _safe_str(row.get("condicion_sanitaria")),
            "tratamiento": {
                "pulverizado": _safe_bool(row.get("tratamiento_pulverizado")),
                "atomizado": _safe_bool(row.get("tratamiento_atomizado")),
                "thermonebulizado": _safe_bool(row.get("tratamiento_thermonebulizado")),
                "nebulizado_ulv": _safe_bool(row.get("tratamiento_nebulizado_ulv")),
                "otros": _safe_str(row.get("tratamiento_otros")),
            },
            "productos": [
                {
                    "producto": _safe_str(row.get(f"producto_{i}_nombre")),
                    "composicion": _safe_str(row.get(f"producto_{i}_composicion")),
                    "lote": _safe_str(row.get(f"producto_{i}_lote")),
                    "fecha_vencimiento": _safe_str(row.get(f"producto_{i}_vencimiento")),
                    "unidad": _safe_str(row.get(f"producto_{i}_unidad")),
                    "concentracion": _safe_str(row.get(f"producto_{i}_concentracion")),
                    "cantidad": _safe_str(row.get(f"producto_{i}_cantidad")),
                }
                for i in range(1, 5)
            ],
            "acciones_correctivas": _safe_str(row.get("acciones_correctivas")),
            "areas_tratadas": _safe_str(row.get("areas_tratadas")),
            "personal_tecnico": [
                _safe_str(row.get("personal_tecnico_1", row.get("personal_tecnico"))),
                _safe_str(row.get("personal_tecnico_2")),
                _safe_str(row.get("personal_tecnico_3")),
                _safe_str(row.get("personal_tecnico_4")),
                _safe_str(row.get("personal_tecnico_5")),
                _safe_str(row.get("personal_tecnico_6")),
            ],
            "hora_inicio": _safe_str(row.get("hora_inicio")),
            "hora_termino": _safe_str(row.get("hora_termino")),
            "numero_certificado": _safe_str(row.get("numero_certificado")),
            "obs_rec": {
                "observacion_a": _safe_str(row.get("observacion_a")),
                "observacion_b": _safe_str(row.get("observacion_b")),
                "observacion_c": _safe_str(row.get("observacion_c")),
                "recomendacion_a": _safe_str(row.get("recomendacion_a")),
                "recomendacion_b": _safe_str(row.get("recomendacion_b")),
                "recomendacion_c": _safe_str(row.get("recomendacion_c")),
            },
            "satisfaccion": _normalize_satisfaccion(row.get("satisfaccion")),
            "status": "draft",
            "last_modified": datetime.now().isoformat(),
        }
    )
    return FichaTecnica.normalize(ficha)


def import_fichas_from_bytes(filename: str, content: bytes) -> list[dict[str, Any]]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        rows = parse_csv_bytes(content)
    elif lower.endswith(".xlsx"):
        rows = parse_xlsx_bytes(content)
    else:
        msg = "Formato no soportado. Use archivos .csv o .xlsx"
        raise ValueError(msg)

    if not rows:
        msg = "El archivo está vacío o no tiene datos válidos"
        raise ValueError(msg)

    return [_row_to_ficha(row, idx + 1) for idx, row in enumerate(rows)]
