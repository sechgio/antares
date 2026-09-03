
from __future__ import annotations

import contextlib
import logging
import re
import sqlite3
import unicodedata
from pathlib import Path
from typing import Any, cast

from backend.core.config_fields import get_field_names, load_fields, sanitize_field_defs, save_fields
from backend.core.exceptions import DatabaseError
from backend.core.repository import _db_lock, _db_read_lock, _db_schema_lock, get_connection, get_read_connection
from backend.utils.paths import user_data_path

logger = logging.getLogger(__name__)

_IDENTIFIER_RE = re.compile(r"^[a-z_][a-z0-9_]*$")


def _validate_identifier(name: str, context: str = "column") -> str:
    if not _IDENTIFIER_RE.match(name):
        msg = f"Invalid SQL {context} name: {name!r}"
        raise ValueError(msg)
    return name


def _qi(name: str) -> str:
    return f'"{name}"'


def _get_connection() -> sqlite3.Connection:
    return get_connection(get_db_path())


def _get_read_connection() -> sqlite3.Connection:
    return get_read_connection(get_db_path())


def _normalize_excel_column_name(name: Any, fallback: str) -> str:
    text = unicodedata.normalize("NFKD", str(name or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-zA-Z0-9_]+", "_", text.strip().lower())
    text = re.sub(r"_+", "_", text).strip("_")
    if not text or not re.match(r"^[a-z_]", text):
        text = fallback
    return text


def _normalize_excel_columns(columns: list[Any]) -> list[str]:
    normalized: list[str] = []
    seen: dict[str, int] = {}
    for idx, column in enumerate(columns, 1):
        base = _normalize_excel_column_name(column, f"columna_{idx}")
        count = seen.get(base, 0)
        seen[base] = count + 1
        normalized.append(base if count == 0 else f"{base}_{count + 1}")
    return normalized


def get_db_path() -> Path:
    return user_data_path("catalogo.db")


def _data_fields(fields: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [f for f in fields if str(f.get("name", "")).lower() != "id"]


def _build_schema(fields: list[dict[str, Any]]) -> str:
    columns = ["id INTEGER PRIMARY KEY AUTOINCREMENT"]
    for f in _data_fields(fields):
        name = _validate_identifier(f["name"])
        if name == "id":
            continue
        quoted_name = _qi(name)
        ftype = f["type"]
        constraints: list[str] = []
        if f.get("required"):
            constraints.append("NOT NULL")
        if f.get("unique"):
            constraints.append("UNIQUE")
        col = f"{quoted_name} {ftype}"
        if constraints:
            col += " " + " ".join(constraints)
        columns.append(col)
    return f"CREATE TABLE IF NOT EXISTS imagenes ({', '.join(columns)})"


def _table_matches_config(cursor: sqlite3.Cursor, fields: list[dict[str, Any]]) -> bool:
    cursor.execute("PRAGMA table_info(imagenes)")
    existing = {row[1]: row[2].upper() for row in cursor.fetchall()}
    expected = {f["name"]: f["type"] for f in fields}
    expected["id"] = "INTEGER"
    return existing == expected


_MIGRATION_NO_OVERLAP_MSG = (
    "Migración abortada: el nuevo esquema no conserva "
    "ninguna columna del catálogo existente y dejaría "
    "la tabla vacía. Conserva al menos una columna "
    "compartida o importa un Excel nuevo."
)


def _schema_diff(
    fields: list[dict[str, Any]],
    existing_cols: dict[str, str],
) -> tuple[dict[str, str], list[str], list[str]]:
    expected_cols = {f["name"]: f["type"] for f in fields}
    expected_cols["id"] = "INTEGER"
    new_cols = {
        name: ftype for name, ftype in expected_cols.items()
        if name not in existing_cols
    }
    removed_cols = [name for name in existing_cols if name not in expected_cols]
    changed_cols = [
        name for name in expected_cols
        if name in existing_cols and existing_cols[name] != expected_cols[name]
    ]
    return new_cols, removed_cols, changed_cols


def _create_indexes(cursor: sqlite3.Cursor, fields: list[dict[str, Any]]) -> None:
    data_fields = _data_fields(fields)
    if not data_fields:
        return

    for f in data_fields:
        name = _validate_identifier(f["name"])
        unique_clause = "UNIQUE" if f.get("unique") else ""
        cursor.execute(
            f"CREATE {unique_clause} INDEX IF NOT EXISTS idx_imagenes_{name} ON imagenes({_qi(name)})"
        )
        cursor.execute(
            f"CREATE INDEX IF NOT EXISTS idx_imagenes_lower_{name} "
            f"ON imagenes(lower({_qi(name)}))"
        )


def init_db(*, allow_catalog_wipe: bool = False) -> None:
    with _db_schema_lock.write():
        _init_db(allow_catalog_wipe=allow_catalog_wipe)


def _init_db(*, allow_catalog_wipe: bool = False) -> None:
    fields = load_fields()
    conn = _get_connection()
    with _db_lock:
        cursor = conn.cursor()

        cursor.execute("BEGIN IMMEDIATE")
        try:
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='imagenes'")
            table_exists = cursor.fetchone() is not None

            if not table_exists:
                cursor.execute(_build_schema(fields))
            elif not _table_matches_config(cursor, fields):
                cursor.execute("PRAGMA table_info(imagenes)")
                existing_cols = {row[1]: row[2].upper() for row in cursor.fetchall()}
                new_cols, removed_cols, changed_cols = _schema_diff(fields, existing_cols)

                if removed_cols or changed_cols:
                    read_cursor: sqlite3.Cursor | None = None
                    try:
                        cursor.execute("SELECT COUNT(*) FROM imagenes")
                        total_old = int(cursor.fetchone()[0])
                        has_rows = total_old > 0
                        old_cols: list[str] = []
                        if has_rows:
                            read_cursor = conn.cursor()
                            read_cursor.execute("SELECT * FROM imagenes")
                            old_cols = [d[0] for d in read_cursor.description]
                    except sqlite3.Error as exc:
                        logger.warning("No se pudieron leer datos antiguos durante migración: %s", exc)
                        if read_cursor is not None:
                            with contextlib.suppress(sqlite3.Error):
                                read_cursor.close()
                            read_cursor = None
                        total_old = 0
                        has_rows = False
                        old_cols = []
                    cursor.execute("ALTER TABLE imagenes RENAME TO imagenes_old")
                    cursor.execute(_build_schema(fields))
                    preserved_cols = [
                        c for c in old_cols
                        if c in {f["name"] for f in fields} and c != "id"
                    ]
                    if has_rows and preserved_cols and read_cursor is not None:
                        data_fields = _data_fields(fields)
                        new_col_names = [_validate_identifier(f["name"]) for f in data_fields]
                        placeholders = ", ".join(["?"] * len(new_col_names))
                        col_names = ", ".join(_qi(c) for c in new_col_names)
                        defaults = {"INTEGER": 0, "REAL": 0.0, "TEXT": "", "BLOB": b""}
                        try:
                            insert_sql = f"INSERT INTO imagenes ({col_names}) VALUES ({placeholders})"
                            chunk: list[list[Any]] = []
                            chunk_size = 500
                            for i, row in enumerate(read_cursor):
                                row_dict = dict(zip(old_cols, row, strict=False))
                                values: list[Any] = []
                                for f in data_fields:
                                    col = f["name"]
                                    if col in row_dict and row_dict[col] is not None:
                                        values.append(row_dict[col])
                                    elif f.get("required"):
                                        default = defaults.get(f["type"], "")
                                        if f.get("unique") and total_old > 1:
                                            if isinstance(default, str):
                                                default = f"{default}{i}" if default else str(i)
                                            elif isinstance(default, (int, float)):
                                                default = default + i
                                            else:
                                                default = str(i)
                                        values.append(default)
                                    else:
                                        values.append(None)
                                chunk.append(values)
                                if len(chunk) >= chunk_size:
                                    cursor.executemany(insert_sql, chunk)
                                    chunk.clear()
                            if chunk:
                                cursor.executemany(insert_sql, chunk)
                            read_cursor.close()
                            read_cursor = None
                            cursor.execute("DROP TABLE imagenes_old")
                        except sqlite3.Error as exc:
                            if read_cursor is not None:
                                read_cursor.close()
                                read_cursor = None
                            logger.error("Fallo migración de datos, se mantiene tabla antigua: %s", exc)
                            cursor.execute("DROP TABLE imagenes")
                            cursor.execute("ALTER TABLE imagenes_old RENAME TO imagenes")
                            raise DatabaseError(f"Migración fallida, esquema anterior preservado: {exc}") from exc
                    else:
                        if read_cursor is not None:
                            read_cursor.close()
                            read_cursor = None
                        if has_rows and not allow_catalog_wipe:
                            cursor.execute("DROP TABLE imagenes")
                            cursor.execute("ALTER TABLE imagenes_old RENAME TO imagenes")
                            raise DatabaseError(_MIGRATION_NO_OVERLAP_MSG)
                        if has_rows:
                            logger.info(
                                "Migración sin solapamiento de columnas (%s → %s): "
                                "catálogo vaciado, no se preservaron filas viejas.",
                                [c for c in old_cols if c != "id"],
                                [f["name"] for f in fields],
                            )
                        cursor.execute("DROP TABLE imagenes_old")
                else:
                    defaults = {"INTEGER": "0", "REAL": "0.0", "TEXT": "''", "BLOB": "NULL"}
                    for col_name, col_type in new_cols.items():
                        safe_name = _validate_identifier(col_name)
                        quoted = _qi(safe_name)
                        default_val = defaults.get(col_type, "''")
                        cursor.execute(f"ALTER TABLE imagenes ADD COLUMN {quoted} {col_type} DEFAULT {default_val}")
                    logger.info("Migración aditiva: se agregaron columnas %s", list(new_cols.keys()))

            _create_indexes(cursor, fields)
            cursor.execute("COMMIT")
        except Exception as exc:
            try:
                if conn.in_transaction:
                    cursor.execute("ROLLBACK")
            except sqlite3.Error as rollback_exc:
                logger.error(
                    "ROLLBACK failed after migration error (%s): %s",
                    type(exc).__name__,
                    rollback_exc,
                )
            raise DatabaseError(f"Inicialización/migración de base de datos fallida: {exc}") from exc


def validate_fields_migration(fields: list[dict[str, Any]]) -> None:
    with _db_schema_lock.read():
        _validate_fields_migration(fields)


def _validate_fields_migration(fields: list[dict[str, Any]]) -> None:
    with _db_schema_lock.read(), _db_read_lock:
        conn = _get_read_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='imagenes'")
        if cursor.fetchone() is None:
            return
        cursor.execute("PRAGMA table_info(imagenes)")
        existing_cols = {row[1]: row[2].upper() for row in cursor.fetchall()}
        _new_cols, removed_cols, changed_cols = _schema_diff(fields, existing_cols)
        if not removed_cols and not changed_cols:
            return
        preserved = [
            c for c in existing_cols
            if c in {f["name"] for f in fields} and c != "id"
        ]
        if preserved:
            return
        cursor.execute("SELECT COUNT(*) FROM imagenes")
        row = cursor.fetchone()
        if not (row and row[0]):
            return
        raise DatabaseError(_MIGRATION_NO_OVERLAP_MSG)


def importar_excel(excel_path: str) -> dict[str, int]:
    try:
        from backend.core.import_guard import serialized_import

        with serialized_import():
            from openpyxl import load_workbook
    except ImportError as exc:
        msg = "openpyxl no está instalado. Ejecuta: pip install openpyxl"
        raise ImportError(msg) from exc

    if not Path(excel_path).exists():
        msg = f"No se encontró el archivo: {excel_path}"
        raise FileNotFoundError(msg)

    wb = load_workbook(excel_path, read_only=True, data_only=False)
    try:
        ws = wb.active
        if ws is None:
            msg = f"El Excel no contiene hojas activas: {excel_path}"
            raise ValueError(msg)
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        columns = _normalize_excel_columns(list(header) if header is not None else [])

        with _db_schema_lock.write(), _db_lock:
            old_fields = load_fields()
            existing_fields = {f["name"]: f for f in old_fields}
            fields = sanitize_field_defs([
                {
                    **existing_fields.get(column, {}),
                    "name": column,
                    "type": existing_fields.get(column, {}).get("type", "TEXT"),
                    "required": bool(existing_fields.get(column, {}).get("required", False)),
                    "unique": bool(existing_fields.get(column, {}).get("unique", False)),
                }
                for column in columns
            ])
            if not fields:
                msg = f"El Excel no contiene columnas válidas para importar: {columns}"
                raise ValueError(msg)

            fields = save_fields(fields)
            try:
                init_db(allow_catalog_wipe=True)
            except Exception:
                with contextlib.suppress(Exception):
                    save_fields(old_fields)
                raise

        with _db_schema_lock.read(), _db_lock:
            if load_fields() != fields:
                raise DatabaseError(
                    "La configuración de campos cambió durante la importación; vuelve a intentarlo."
                )
            field_names = [_validate_identifier(f["name"]) for f in fields]
            required = [f["name"] for f in fields if f.get("required")]

            conn = _get_connection()
            cursor = conn.cursor()

            try:
                cursor.execute("BEGIN")

                cursor.execute("DELETE FROM imagenes")

                placeholders = ", ".join(["?"] * len(field_names))
                col_names = ", ".join(_qi(fn) for fn in field_names)
                sql = f"INSERT INTO imagenes ({col_names}) VALUES ({placeholders})"

                chunk: list[list[Any]] = []
                chunk_size = 500
                inserted = 0
                skipped = 0
                required_set = set(required)
                for row in rows_iter:
                    row_dict = dict(zip(columns, row, strict=False))
                    values: list[Any] = []
                    valid = True
                    for fn in field_names:
                        val = row_dict.get(fn)
                        if val is not None and str(val).strip():
                            values.append(str(val).strip())
                        elif fn in required_set:
                            valid = False
                            break
                        else:
                            values.append(None)
                    if valid:
                        chunk.append(values)
                        if len(chunk) >= chunk_size:
                            cursor.executemany(sql, chunk)
                            inserted += len(chunk)
                            chunk.clear()
                    else:
                        skipped += 1

                if chunk:
                    cursor.executemany(sql, chunk)
                    inserted += len(chunk)

                cursor.execute("COMMIT")
                try:
                    conn.execute("ANALYZE imagenes")
                except sqlite3.Error:
                    logger.debug("ANALYZE after import failed", exc_info=True)
                return {"inserted": inserted, "skipped": skipped}
            except sqlite3.Error as exc:
                cursor.execute("ROLLBACK")
                msg = f"Error importando datos: {exc}"
                raise DatabaseError(msg) from exc
    finally:
        wb.close()


def exportar_excel(excel_path: str) -> int:
    try:
        import pandas as pd
    except ImportError as exc:
        msg = "pandas no está instalado."
        raise ImportError(msg) from exc

    with _db_schema_lock.read(), _db_read_lock:
        conn = _get_read_connection()
        field_names = [_validate_identifier(fn) for fn in get_field_names()]
        cols = ", ".join(_qi(fn) for fn in field_names)
        df = pd.read_sql_query(f"SELECT {cols} FROM imagenes ORDER BY id", conn)

    df.to_excel(excel_path, index=False, engine="openpyxl")
    return len(df)


def _record_code_match(
    result: dict[str, dict[str, Any]],
    code_rowids: dict[str, int],
    val: str,
    row_id: int,
    row_dict: dict[str, Any],
) -> None:
    prev_id = code_rowids.get(val)
    if prev_id is not None and prev_id != row_id:
        logger.warning(
            "Código duplicado en catálogo: %r coincide con varios "
            "registros (rowid %s vs %s). Se conserva el primero.",
            val, prev_id, row_id,
        )
        return
    code_rowids[val] = row_id
    result[val] = row_dict


def buscar_lote_por_codigos(codigos: list[str]) -> dict[str, dict[str, Any]]:
    if not codigos:
        return {}
    with _db_schema_lock.read(), _db_read_lock:
        conn = _get_read_connection()
        cursor = conn.cursor()
        field_names = [_validate_identifier(fn) for fn in get_field_names()]
        if not field_names:
            return {}

        query_by_fold: dict[str, str] = {}
        for raw in codigos:
            text = str(raw).strip()
            if text:
                query_by_fold.setdefault(text.casefold(), text)
        if not query_by_fold:
            return {}
        folded_codes = list(query_by_fold.keys())

        result: dict[str, dict[str, Any]] = {}
        code_rowids: dict[str, int] = {}
        cols = ", ".join(_qi(fn) for fn in field_names)

        preferred = "codigo" if "codigo" in field_names else field_names[0]

        def _query_key_for(val: str) -> str | None:
            return query_by_fold.get(val.casefold())

        def _scan_column(codes: list[str], column: str) -> None:
            CHUNK = 900
            for i in range(0, len(codes), CHUNK):
                chunk = codes[i:i + CHUNK]
                placeholders = ", ".join(["?"] * len(chunk))
                cursor.execute(
                    f"SELECT rowid AS __antares_rowid__, {cols} FROM imagenes "
                    f"WHERE lower({_qi(column)}) IN ({placeholders})",
                    chunk,
                )
                for row in cursor.fetchall():
                    row_dict = dict(row)
                    row_id = row_dict.pop("__antares_rowid__")
                    for fn in field_names:
                        val = str(row_dict.get(fn, "") or "").strip()
                        query_key = _query_key_for(val) if val else None
                        if query_key:
                            _record_code_match(result, code_rowids, query_key, row_id, row_dict)

        _scan_column(folded_codes, preferred)

        unresolved = [c for c in folded_codes if query_by_fold[c] not in result]
        if unresolved and len(field_names) > 1:
            CHUNK = max(1, 900 // len(field_names))
            other_fields = [fn for fn in field_names if fn != preferred]
            for i in range(0, len(unresolved), CHUNK):
                chunk = unresolved[i:i + CHUNK]
                placeholders = ", ".join(["?"] * len(chunk))
                conditions = " OR ".join(
                    [f"lower({_qi(fn)}) IN ({placeholders})" for fn in other_fields]
                )
                params = chunk * len(other_fields)
                cursor.execute(
                    f"SELECT rowid AS __antares_rowid__, {cols} FROM imagenes WHERE {conditions}",
                    params,
                )
                for row in cursor.fetchall():
                    row_dict = dict(row)
                    row_id = row_dict.pop("__antares_rowid__")
                    for fn in field_names:
                        val = str(row_dict.get(fn, "") or "").strip()
                        query_key = _query_key_for(val) if val else None
                        if query_key:
                            _record_code_match(result, code_rowids, query_key, row_id, row_dict)
        return result


def buscar_por_columna(codigos: list[str], column: str) -> dict[str, dict[str, Any]]:
    if not codigos or not column:
        return {}
    safe_column = _validate_identifier(column)
    with _db_schema_lock.read(), _db_read_lock:
        conn = _get_read_connection()
        cursor = conn.cursor()
        field_names_list = [_validate_identifier(fn) for fn in get_field_names()]
        field_names = set(field_names_list)
        if safe_column not in field_names:
            return {}

        query_by_fold: dict[str, str] = {}
        for raw in codigos:
            text = str(raw).strip()
            if text:
                query_by_fold.setdefault(text.casefold(), text)
        if not query_by_fold:
            return {}
        folded_codes = list(query_by_fold.keys())

        result: dict[str, dict[str, Any]] = {}
        code_rowids: dict[str, int] = {}
        CHUNK = 900
        cols = ", ".join(_qi(fn) for fn in field_names_list)
        for i in range(0, len(folded_codes), CHUNK):
            chunk = folded_codes[i:i + CHUNK]
            placeholders = ", ".join(["?"] * len(chunk))
            cursor.execute(
                f"SELECT rowid AS __antares_rowid__, {cols} FROM imagenes "
                f"WHERE lower({_qi(safe_column)}) IN ({placeholders})",
                chunk,
            )
            for row in cursor.fetchall():
                row_dict = dict(row)
                row_id = row_dict.pop("__antares_rowid__")
                val = str(row_dict.get(safe_column, "") or "").strip()
                query_key = query_by_fold.get(val.casefold()) if val else None
                if query_key:
                    _record_code_match(result, code_rowids, query_key, row_id, row_dict)
        return result


def obtener_todos(limit: int | None = None, offset: int = 0) -> list[dict[str, Any]]:
    with _db_schema_lock.read(), _db_read_lock:
        conn = _get_read_connection()
        cursor = conn.cursor()
        field_names = [_validate_identifier(fn) for fn in get_field_names()]
        cols = ", ".join(_qi(fn) for fn in field_names)
        sql = f"SELECT {cols} FROM imagenes ORDER BY id"
        params: list[Any] = []
        if limit is not None:
            sql += " LIMIT ? OFFSET ?"
            params = [limit, offset]
        cursor.execute(sql, params)
        rows = cursor.fetchall()
    return [dict(r) for r in rows]


def limpiar_base_datos() -> int:
    db_path = get_db_path()
    if not db_path.exists():
        return 0
    with _db_lock:
        conn = _get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM imagenes")
        row = cursor.fetchone()
        count = int(row[0]) if row else 0
        cursor.execute("DELETE FROM imagenes")
    return count


_MAPPING_ID_ALIASES = ("id", "codigo", "code", "filename", "archivo", "nombre original")
_MAPPING_RENAME_ALIASES = ("renombre", "rename", "new_name", "newname", "nombre nuevo", "nuevo_nombre", "nuevonombre")


def _normalize_header_alias(header: str) -> str:
    text = str(header).lower().strip()
    text = "".join(c for c in unicodedata.normalize("NFKD", text) if unicodedata.category(c) != "Mn")
    text = re.sub(r"[\s_\-]+", " ", text)
    text = " ".join(text.split())
    return text


def _detect_column(columns: list[str], aliases: tuple[str, ...]) -> str | None:
    for alias in aliases:
        normalized_alias = _normalize_header_alias(alias)
        for col in columns:
            if _normalize_header_alias(col) == normalized_alias:
                return col
    return None


def parse_id_rename_mapping(
    excel_path: str,
    id_column: str | None = None,
    rename_column: str | None = None,
) -> dict[str, str]:
    result = parse_id_rename_mapping_full(excel_path, id_column, rename_column)
    return cast(dict[str, str], result["mapping"])


def parse_id_rename_mapping_full(
    excel_path: str,
    id_column: str | None = None,
    rename_column: str | None = None,
) -> dict[str, Any]:
    try:
        from backend.core.import_guard import serialized_import

        with serialized_import():
            import pandas as pd
    except ImportError as exc:
        msg = "pandas no está instalado. Ejecuta: pip install pandas openpyxl"
        raise ImportError(msg) from exc

    from backend.utils.validators import sanitizar_nombre

    if not Path(excel_path).exists():
        msg = f"No se encontró el archivo: {excel_path}"
        raise FileNotFoundError(msg)

    df = pd.read_excel(
        excel_path,
        dtype=str,
        engine="openpyxl",
        engine_kwargs={"read_only": True},
    )
    df.columns = _normalize_excel_columns(list(df.columns))
    columns = list(df.columns)

    if len(columns) < 2:
        msg = (
            f"El Excel de mapeo necesita al menos 2 columnas (una ID y una de nuevo nombre). "
            f"Columnas encontradas: {columns}"
        )
        raise ValueError(msg)

    raw_id_column = id_column or _detect_column(columns, _MAPPING_ID_ALIASES)
    raw_rename_column = rename_column or _detect_column(columns, _MAPPING_RENAME_ALIASES)

    chosen_id = raw_id_column if raw_id_column in df.columns else None
    chosen_rename = raw_rename_column if raw_rename_column in df.columns else None

    if not chosen_id:
        msg = (
            f"No se detectó una columna ID. Usa una columna como ID, Código, Code, Filename, Archivo, "
            f"o indica id_column. Columnas disponibles: {columns}"
        )
        raise ValueError(msg)
    if not chosen_rename:
        msg = (
            f"No se detectó una columna de nuevo nombre. Usa RENOMBRE, Nombre, New_Name, etc., "
            f"o indica rename_column. Columnas disponibles: {columns}"
        )
        raise ValueError(msg)
    if chosen_id == chosen_rename:
        msg = "La columna ID y la columna de nuevo nombre no pueden ser la misma."
        raise ValueError(msg)

    result: dict[str, str] = {}
    seen_ids: set[str] = set()

    for row_idx, row in df.iterrows():
        excel_row = int(row_idx) + 2
        raw_id = row.get(chosen_id)
        raw_rename = row.get(chosen_rename)

        if pd.isna(raw_id) or not str(raw_id).strip():
            msg = f"ID vacío en la fila {excel_row} de la columna '{chosen_id}'"
            raise ValueError(msg)
        if pd.isna(raw_rename) or not str(raw_rename).strip():
            msg = f"Nuevo nombre vacío en la fila {excel_row} de la columna '{chosen_rename}'"
            raise ValueError(msg)

        id_key = str(raw_id).strip()
        if id_key in seen_ids:
            msg = f"ID duplicado '{id_key}' en la fila {excel_row} de la columna '{chosen_id}'"
            raise ValueError(msg)
        seen_ids.add(id_key)

        rename_val = sanitizar_nombre(str(raw_rename).strip())
        if not rename_val:
            msg = f"Nuevo nombre inválido en la fila {excel_row} de la columna '{chosen_rename}'"
            raise ValueError(msg)
        result[id_key] = rename_val

    return {
        "mapping": result,
        "id_column": chosen_id,
        "rename_column": chosen_rename,
        "columns": columns,
    }


def generar_plantilla_excel(ruta_salida: str) -> int:
    try:
        import pandas as pd
    except ImportError as exc:
        msg = "pandas no está instalado."
        raise ImportError(msg) from exc

    fields = load_fields()
    columns = [f["name"] for f in fields]
    df = pd.DataFrame(columns=columns)
    sample: list[str] = []
    for f in fields:
        fname: str = f["name"]
        if fname == "codigo":
            sample.append("IMG-001")
        elif fname == "nombre":
            sample.append("Producto Ejemplo")
        elif fname == "categoria":
            sample.append("Categoria A")
        elif fname == "marca":
            sample.append("Marca X")
        elif fname == "modelo":
            sample.append("Modelo 2024")
        elif fname == "descripcion":
            sample.append("Descripción de prueba")
        else:
            sample.append(f"Ejemplo {fname}")
    df.loc[0] = sample
    df.to_excel(ruta_salida, index=False, engine="openpyxl")
    return len(df)
