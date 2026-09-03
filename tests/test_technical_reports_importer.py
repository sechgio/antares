from io import BytesIO

from openpyxl import Workbook

from backend.core.technical_reports.importer import import_reports_from_bytes, normalize_header_value


def test_csv_semicolon_import_maps_human_headers() -> None:
    content = (
        b"Nro Informe;Centro de Servicio;Codigo Infraestructura;Tipo;Volumen;Caja Registro;Mes\n"
        b"3;SUR;RES-01;ELEVADO;150;X;5\n"
    )

    reports = import_reports_from_bytes("datos.csv", content)

    assert reports[0]["id"] == "RPT-0003"
    assert reports[0]["header"]["cs"] == "SUR"
    assert reports[0]["metadata"]["mes"] == "MAYO"
    assert reports[0]["inspeccion"]["caja_registro"] == "normal"


def test_csv_import_assigns_unique_ids_when_missing_report_number() -> None:
    content = (
        b"Centro de Servicio;Codigo Infraestructura;Tipo\n"
        b"SUR;RES-01;ELEVADO\n"
        b"NORTE;CIS-02;CISTERNA\n"
    )

    reports = import_reports_from_bytes("datos.csv", content)

    assert [report["id"] for report in reports] == ["RPT-0001", "RPT-0002"]
    assert [report["metadata"]["informe_id"] for report in reports] == [1, 2]


def test_xlsx_import_reads_rows() -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Informe", "CS", "Codigo", "Tipo", "Volumen", "Descarga"])
    ws.append([4, "NORTE", "CIS-02", "CISTERNA", 80, "MALO"])
    buf = BytesIO()
    wb.save(buf)

    reports = import_reports_from_bytes("datos.xlsx", buf.getvalue())

    assert reports[0]["id"] == "RPT-0004"
    assert reports[0]["header"]["tipo"] == "CISTERNA"
    assert reports[0]["inspeccion"]["descarga"] == "critico"


def test_xlsx_import_maps_reference_human_header_aliases() -> None:
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Informe",
        "Numero Suministro",
        "Marco y Tapa Sanitaria",
        "Observaciones Marco y Tapa Sanitaria",
        "Sugerencias Marco y Tapa Sanitaria",
        "Valv Cond 2",
        "Canastillas Aduccion 12",
        "Observaciones Canastilla Succion",
        "Sugerencias Succion",
    ])
    ws.append([8, "NIS-99", "MALO", "CORROIDA", "CAMBIAR", 2, 1, "CON OXIDO", "MANTENIMIENTO"])
    buf = BytesIO()
    wb.save(buf)

    reports = import_reports_from_bytes("referencia.xlsx", buf.getvalue())

    report = reports[0]
    assert report["header"]["suministro"] == "NIS-99"
    assert report["inspeccion"]["marco_tapa"] == "critico"
    assert report["inspeccion"]["observaciones_marco_tapa"] == "CORROIDA"
    assert report["inspeccion"]["sugerencias_marco_tapa"] == "CAMBIAR"
    assert report["valvulas"]["diametros"]["2"] == 2
    assert report["canastillas"]["aduccion"]["14"] == 1
    assert report["canastillas"]["observaciones_succion"] == "CON OXIDO"
    assert report["canastillas"]["sugerencias_succion"] == "MANTENIMIENTO"


def test_xlsx_import_maps_recommendation_headers_to_sugerencias() -> None:
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Informe",
        "Recomendaciones Marco y Tapa Sanitaria",
        "Recomendaciones Valvulas Desague",
        "Recomendaciones Canastillas Succion",
        "Recomendaciones",
    ])
    ws.append([9, "CAMBIAR TAPA", "MANTENIMIENTO", "INSTALAR", "REVISAR EN CAMPO"])
    buf = BytesIO()
    wb.save(buf)

    reports = import_reports_from_bytes("recomendaciones.xlsx", buf.getvalue())

    report = reports[0]
    assert report["inspeccion"]["sugerencias_marco_tapa"] == "CAMBIAR TAPA"
    assert report["valvulas"]["sugerencias_desague"] == "MANTENIMIENTO"
    assert report["canastillas"]["sugerencias_succion"] == "INSTALAR"
    assert report["sugerencias"] == "REVISAR EN CAMPO"


def test_normalize_header_value_removes_accents_and_separators() -> None:
    assert normalize_header_value("Código de Infraestructura") == "codigoinfraestructura"


def test_csv_import_maps_sgio_column() -> None:
    content = (
        b"Informe;Contratista;SGIO;Codigo Infraestructura;Tipo\n"
        b"5;ACCIONA;454654001;RES-05;ELEVADO\n"
    )

    reports = import_reports_from_bytes("datos.csv", content)

    assert reports[0]["header"]["contratista"] == "ACCIONA"
    assert reports[0]["header"]["sgio"] == "454654001"


def test_csv_import_handles_latin1_encoding() -> None:
    content = "N° Informe;Centro de Servicio;Ubicación;Contratista\n1;Zonal Cañete;Concepción N° 123;Compañía S.A.".encode("latin-1")

    reports = import_reports_from_bytes("latin1_data.csv", content)

    assert reports[0]["id"] == "RPT-0001"
    assert reports[0]["header"]["cs"] == "Zonal Cañete"
    assert reports[0]["header"]["ubicacion"] == "Concepción N° 123"
    assert reports[0]["header"]["contratista"] == "Compañía S.A."


def test_xlsx_import_handles_datetime_cells_and_title_row() -> None:
    from datetime import date, datetime
    wb = Workbook()
    ws = wb.active
    ws.append(["INFORMES TÉCNICOS 2026"])
    ws.append(["N°", "CS", "Fecha", "Caja Registro", "Marco Tapa"])
    ws.append([10, "SUR", datetime(2026, 5, 15), "CONFORME", "INOPERATIVO"])
    ws.append([11, "NORTE", date(2026, 8, 20), "OPERATIVO", "MALO"])
    buf = BytesIO()
    wb.save(buf)

    reports = import_reports_from_bytes("reportes_con_fechas.xlsx", buf.getvalue())

    assert len(reports) == 2
    assert reports[0]["id"] == "RPT-0010"
    assert reports[0]["metadata"]["dia"] == 15
    assert reports[0]["metadata"]["mes"] == "MAYO"
    assert reports[0]["metadata"]["anio"] == 2026
    assert reports[0]["inspeccion"]["caja_registro"] == "normal"
    assert reports[0]["inspeccion"]["marco_tapa"] == "critico"

    assert reports[1]["id"] == "RPT-0011"
    assert reports[1]["metadata"]["dia"] == 20
    assert reports[1]["metadata"]["mes"] == "AGOSTO"
    assert reports[1]["metadata"]["anio"] == 2026
    assert reports[1]["inspeccion"]["caja_registro"] == "normal"
    assert reports[1]["inspeccion"]["marco_tapa"] == "critico"


def test_single_fecha_column_string_parsing() -> None:
    content = (
        b"Nro;CS;Fecha;Tipo\n"
        b"12;SUR;25/07/2026;ELEVADO\n"
    )

    reports = import_reports_from_bytes("fecha_string.csv", content)

    assert reports[0]["metadata"]["dia"] == 25
    assert reports[0]["metadata"]["mes"] == "JULIO"
    assert reports[0]["metadata"]["anio"] == 2026

