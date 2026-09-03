from io import BytesIO

from openpyxl import Workbook

from backend.core.informes_v2.importer import (
    TEMPLATE_HEADERS,
    import_reports_from_bytes,
    normalize_header_value,
)
from backend.core.informes_v2.template_xlsx import build_template_xlsx_bytes


def test_normalize_header_strips_noise() -> None:
    assert normalize_header_value("Fecha de Ejecución") == "fechaejecucion"
    assert normalize_header_value("Valv Cond 2\"") == "valvcond2"


def test_csv_import_maps_headers_and_valves() -> None:
    content = (
        b"ID;Informe;Estacion;Tipo;Volumen;Distrito;Suministro;SGIO;Valv Cond 2;Valv Cond Oper;Lin Aduccion 4;Largo;Observacion\n"
        b"R-900;1;R 900 Elevado;ELEVADO;900;Villa El Salvador;2748175;SG-1;2;1;3;4.5;Sin obs\n"
    )
    reports = import_reports_from_bytes("datos.csv", content)
    report = reports[0]
    assert report["id"] == "IV2-0001"
    assert report["header"]["photo_id"] == "R-900"
    assert report["header"]["estacion"] == "R 900 Elevado"
    assert report["header"]["volumen"] == 900
    assert report["valvulas"]["conduccion"]["diametros"]["2"] == 2
    assert report["valvulas"]["conduccion"]["oper"] == 1
    assert report["linea"]["aduccion"]["diametros"]["4"] == 3
    assert report["medidas"]["largo"] == "4.5"
    assert report["medidas"]["observacion"] == "Sin obs"


def test_xlsx_import_assigns_ids_when_missing() -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Estacion", "Tipo", "Suministro"])
    ws.append(["Est A", "CISTERNA", "111"])
    ws.append(["Est B", "ELEVADO", "222"])
    buf = BytesIO()
    wb.save(buf)

    reports = import_reports_from_bytes("datos.xlsx", buf.getvalue())
    assert [r["id"] for r in reports] == ["IV2-0001", "IV2-0002"]
    assert reports[0]["header"]["photo_id"] == "Est A"
    assert reports[1]["header"]["tipo"] == "ELEVADO"


def test_template_xlsx_has_canonical_headers() -> None:
    content = build_template_xlsx_bytes()
    from openpyxl import load_workbook

    loaded = load_workbook(BytesIO(content))
    headers = [cell.value for cell in next(loaded.active.iter_rows(min_row=1, max_row=1))]
    assert headers == TEMPLATE_HEADERS
    assert "ID" in headers
    assert "Valv Cond 2" in headers
    assert "Lin Impulsion 16" in headers
    assert "Tirante Limpieza" in headers


def test_all_template_headers_are_mapped() -> None:
    from backend.core.informes_v2.importer import COLUMN_MAPPING

    unmapped = [
        header
        for header in TEMPLATE_HEADERS
        if normalize_header_value(header) not in COLUMN_MAPPING
    ]
    assert unmapped == [], f"Headers de plantilla sin mapeo: {unmapped}"


def test_template_xlsx_roundtrip_maps_all_sections() -> None:
    from openpyxl import load_workbook

    content = build_template_xlsx_bytes()
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    by_header = {h: i + 1 for i, h in enumerate(headers)}

    row_values = {
        "ID": "7098",
        "Informe": 3,
        "Estacion": "R 900 Elevado",
        "Tipo": "ELEVADO",
        "Volumen": 900,
        "Ubicacion": "Av. Principal 123",
        "Distrito": "Villa El Salvador",
        "Fecha Ejecucion": "2026-03-15",
        "Suministro": "2748175",
        "SGIO": "SG-9",
        "Valv Cond 2": 2,
        "Valv Cond Oper": 1,
        "Valv Cond No Op": 0,
        "Valv Cond Obs": "OK cond",
        "Valv Imp 4": 1,
        "Lin Aduccion 6": 3,
        "Lin Impulsion 8": 4,
        "Lin Impulsion Oper": 2,
        "Lin Impulsion Obs": "rebombeo ok",
        "Lin Rebose 10": 1,
        "Largo": "5.5",
        "Ancho": "3.2",
        "Diametro": "2.1",
        "Altura Rebose": "1.0",
        "Altura Total": "4.0",
        "Tirante Limpieza": "0.3",
        "Observacion": "Sin hallazgos",
    }
    for header, value in row_values.items():
        ws.cell(row=2, column=by_header[header], value=value)

    buf = BytesIO()
    wb.save(buf)
    reports = import_reports_from_bytes("plantilla.xlsx", buf.getvalue())
    assert len(reports) == 1
    report = reports[0]

    assert report["id"] == "IV2-0003"
    assert report["header"]["photo_id"] == "7098"
    assert report["header"]["estacion"] == "R 900 Elevado"
    assert report["header"]["tipo"] == "ELEVADO"
    assert report["header"]["volumen"] == 900
    assert report["header"]["ubicacion"] == "Av. Principal 123"
    assert report["header"]["distrito"] == "Villa El Salvador"
    assert report["header"]["fecha_ejecucion"] == "15/03/2026"
    assert report["header"]["suministro"] == "2748175"
    assert report["header"]["sgio"] == "SG-9"

    assert report["valvulas"]["conduccion"]["diametros"]["2"] == 2
    assert report["valvulas"]["conduccion"]["oper"] == 1
    assert report["valvulas"]["conduccion"]["observaciones"] == "OK cond"
    assert report["valvulas"]["impulsion"]["diametros"]["4"] == 1

    assert report["linea"]["aduccion"]["diametros"]["6"] == 3
    assert report["linea"]["impulsion_rebombeo"]["diametros"]["8"] == 4
    assert report["linea"]["impulsion_rebombeo"]["oper"] == 2
    assert report["linea"]["impulsion_rebombeo"]["observaciones"] == "rebombeo ok"
    assert report["linea"]["rebose"]["diametros"]["10"] == 1

    assert report["medidas"]["largo"] == "5.5"
    assert report["medidas"]["ancho"] == "3.2"
    assert report["medidas"]["diametro"] == "2.1"
    assert report["medidas"]["altura_rebose"] == "1.0"
    assert report["medidas"]["altura_total"] == "4.0"
    assert report["medidas"]["tirante_limpieza"] == "0.3"
    assert report["medidas"]["observacion"] == "Sin hallazgos"
