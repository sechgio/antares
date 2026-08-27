"""Tests para el provider de mapas estáticos (reemplazo de Playwright)."""

import math
import os
import tempfile
import urllib.parse
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

import pandas as pd
import pytest
from PIL import Image
from pypdf import PdfReader

from backend.handlers import ubicaciones as ub


def _png_bytes(size: tuple[int, int], color: tuple[int, int, int] = (70, 130, 180)) -> bytes:
    buf = BytesIO()
    Image.new("RGB", size, color).save(buf, format="PNG")
    return buf.getvalue()


class TestProviderResolution:
    def test_default_is_osm(self, monkeypatch: pytest.MonkeyPatch) -> None:
        monkeypatch.delenv("ANTARES_MAP_PROVIDER", raising=False)
        assert ub._resolve_provider(None) == "osm"

    def test_env_overrides_default(self, monkeypatch: pytest.MonkeyPatch) -> None:
        monkeypatch.setenv("ANTARES_MAP_PROVIDER", "google")
        assert ub._resolve_provider(None) == "google"

    def test_payload_overrides_env(self, monkeypatch: pytest.MonkeyPatch) -> None:
        monkeypatch.setenv("ANTARES_MAP_PROVIDER", "google")
        assert ub._resolve_provider({"provider": "osm"}) == "osm"

    def test_lowercased(self) -> None:
        assert ub._resolve_provider({"provider": "GOOGLE"}) == "google"


def test_resolve_api_key(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.delenv("ANTARES_GOOGLE_MAPS_KEY", raising=False)
    monkeypatch.delenv("ANTARES_MAPS_API_KEY", raising=False)
    assert ub._resolve_api_key(None) is None
    assert ub._resolve_api_key({"google_maps_key": "abc"}) == "abc"
    assert ub._resolve_api_key({"api_key": "xyz"}) == "xyz"
    monkeypatch.setenv("ANTARES_GOOGLE_MAPS_KEY", "envkey")
    assert ub._resolve_api_key(None) == "envkey"
    # payload wins over env
    assert ub._resolve_api_key({"api_key": "pk"}) == "pk"


class TestCapFetchSize:
    def test_no_upscale_when_under_cap(self) -> None:
        assert ub._cap_fetch_size(600, 800) == (600, 800)

    def test_caps_long_side_preserving_aspect(self) -> None:
        w, h = ub._cap_fetch_size(2480, 3386)
        assert max(w, h) <= ub._MAP_FETCH_MAX_DIM
        # aspect preserved within rounding
        assert w / h == pytest.approx(2480 / 3386, rel=0.01)

    def test_never_zero(self) -> None:
        w, h = ub._cap_fetch_size(1, 1)
        assert w >= 1 and h >= 1


def test_lonlat_to_webmercator_pixel_origin() -> None:
    # Web Mercator: (lon=-180, lat~0) maps to x=0; lat=0 maps to y = n*128.
    x, y = ub._lonlat_to_webmercator_pixel(-180.0, 0.0, 0)
    assert x == pytest.approx(0.0, abs=1e-3)
    assert y == pytest.approx(128.0, abs=1e-3)


@pytest.mark.parametrize(
    ("lat", "lon"),
    [
        (-12.0464, -77.0428),  # Plaza de Armas de Lima
        (40.6892, -74.0445),   # Estatua de la Libertad
    ],
)
def test_webmercator_center_round_trips_to_requested_coordinates(lat: float, lon: float) -> None:
    zoom = 18
    center_x, center_y = ub._lonlat_to_webmercator_pixel(lon, lat, zoom)
    world_size = (2 ** zoom) * ub._OSM_TILE_SIZE

    recovered_lon = center_x / world_size * 360.0 - 180.0
    recovered_lat = math.degrees(math.atan(math.sinh(math.pi * (1.0 - 2.0 * center_y / world_size))))

    assert recovered_lat == pytest.approx(lat, abs=1e-5)
    assert recovered_lon == pytest.approx(lon, abs=1e-5)


def test_fetch_static_map_osm_returns_image(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("ANTARES_MAP_PROVIDER", "osm")
    tile = _png_bytes((ub._OSM_TILE_SIZE, ub._OSM_TILE_SIZE), (60, 120, 160))
    monkeypatch.setattr(ub, "_http_get", lambda url, headers, timeout=ub._HTTP_TIMEOUT: tile)

    data = ub.fetch_static_map(-12.046, -77.042, 800, 600, zoom=18, provider="osm")
    img = Image.open(BytesIO(data))
    assert img.size == ub._cap_fetch_size(800, 600)
    # A real (colored) map passes the tiles heuristic.
    assert ub._screenshot_has_map_tiles(data)


def test_osm_live_smoke_plaza_de_armas_lima(tmp_path: Path) -> None:
    if os.environ.get("ANTARES_RUN_NETWORK_TESTS") != "1":
        pytest.skip("set ANTARES_RUN_NETWORK_TESTS=1 to run live OSM smoke test")

    data = ub.fetch_static_map(-12.0464, -77.0428, 600, 400, zoom=18, provider="osm")
    output = tmp_path / "plaza-de-armas-lima.png"
    output.write_bytes(data)

    with Image.open(output) as image:
        assert image.size == (600, 400)
    assert ub._screenshot_has_map_tiles(data)


def test_fetch_xyz_tiles_map_downloads_tiles_in_parallel(monkeypatch: pytest.MonkeyPatch) -> None:
    tile = _png_bytes((ub._OSM_TILE_SIZE, ub._OSM_TILE_SIZE), (60, 120, 160))
    calls: list[str] = []

    def fake_get(url: str, headers: dict[str, str], timeout: int = ub._HTTP_TIMEOUT) -> bytes:
        calls.append(url)
        return tile

    monkeypatch.setattr(ub, "_http_get", fake_get)
    img = ub._fetch_xyz_tiles_map(-12.046, -77.042, 850, 568, 18, ub._XYZ_PROVIDERS["osm"])
    assert img.size == (850, 568)
    assert len(calls) >= 2


def test_fetch_static_map_fallback_on_http_failure(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(ub, "_http_get", lambda *a, **k: None)
    data = ub.fetch_static_map(-12.0, -77.0, 800, 600, zoom=18, provider="osm")
    img = Image.open(BytesIO(data))
    fw, fh = ub._cap_fetch_size(800, 600)
    assert img.size == (fw, fh)
    # Uniform gray placeholder does NOT pass the tiles heuristic.
    assert not ub._screenshot_has_map_tiles(data)


def test_fetch_static_map_google_without_key_falls_back(monkeypatch: pytest.MonkeyPatch) -> None:
    data = ub.fetch_static_map(-12.0, -77.0, 800, 600, zoom=18, provider="google", api_key=None)
    assert not ub._screenshot_has_map_tiles(data)


def test_fetch_static_map_google_with_key(monkeypatch: pytest.MonkeyPatch) -> None:
    called: dict[str, str] = {}

    def fake_get(url, headers, timeout=ub._HTTP_TIMEOUT):
        called["url"] = url
        return _png_bytes((640, 640), (90, 140, 190))

    monkeypatch.setattr(ub, "_http_get", fake_get)
    data = ub.fetch_static_map(-12.0, -77.0, 800, 600, zoom=18, provider="google", api_key="TESTKEY")
    parsed = urllib.parse.urlparse(called["url"])
    query = urllib.parse.parse_qs(parsed.query)
    assert parsed.netloc == "maps.googleapis.com"
    assert query["center"] == ["-12.0,-77.0"]
    assert query["key"] == ["TESTKEY"]
    assert ub._screenshot_has_map_tiles(data)


def test_google_static_map_size_preserves_aspect_ratio() -> None:
    assert ub._google_static_map_size(800, 600) == (640, 480)
    assert ub._google_static_map_size(600, 800) == (480, 640)
    assert ub._google_static_map_size(600, 400) == (600, 400)


def test_fetch_static_map_google_requests_proportional_viewport(monkeypatch: pytest.MonkeyPatch) -> None:
    called: dict[str, str] = {}

    def fake_get(url, headers, timeout=ub._HTTP_TIMEOUT):
        called["url"] = url
        return _png_bytes((640, 640), (90, 140, 190))

    monkeypatch.setattr(ub, "_http_get", fake_get)
    ub.fetch_static_map(-12.0, -77.0, 749, 1024, zoom=18, provider="google", api_key="TESTKEY")

    query = urllib.parse.parse_qs(urllib.parse.urlparse(called["url"]).query)
    assert query["size"] == ["468x640"]


def test_handle_generar_ubicaciones_skips_non_numeric_coords(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    """Una fila con lat/lon no numérico debe skiparse, no crashar el batch.

    El filtro histórico sólo rechazaba NaN (``pd.isna("abc")`` es False), así que
    un texto en la columna de coordenadas pasaba y luego ``float(datos['lat'])``
    levantaba ValueError dentro del worker, abortando todo el batch vía ex.map.
    """
    import openpyxl

    monkeypatch.setattr(ub, "fetch_static_map", lambda *a, **k: _png_bytes((256, 256)))

    xlsx = tmp_path / "coords.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["latitud", "longitud"])
    ws.append([-12.0, -77.0])   # válida
    ws.append(["abc", -77.0])   # lat no-numérica -> debe skiparse
    wb.save(xlsx)

    out_dir = tmp_path / "out"
    result = ub.handle_generar_ubicaciones({
        "excelPath": str(xlsx),
        "outputDir": str(out_dir),
        "formato": "vertical",
        "consolidado": False,
    })

    assert result["generados"] == 1
    pdfs = list(out_dir.glob("*.pdf"))
    assert len(pdfs) == 1


def test_handle_generar_ubicaciones_continues_after_row_failure(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    """Una fila que falla al renderizar (IO, imagen corrupta, etc.) no debe
    abortar el batch: las demás filas se procesan y se reporta el conteo de
    fallidos. Antes el try/finally sin except dejaba que ex.map re-lanzara y
    el handler devolvía success:False con archivos huérfanos."""
    import openpyxl

    monkeypatch.setattr(ub, "fetch_static_map", lambda *a, **k: _png_bytes((256, 256)))

    xlsx = tmp_path / "coords.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["latitud", "longitud"])
    ws.append([-12.0, -77.0])   # ID-1: ok
    ws.append([-13.0, -78.0])   # ID-2: simulamos fallo de render
    wb.save(xlsx)

    real_gen = ub.generar_imagen_ubicacion

    def flaky_gen(d, out_path, formato, map_opts=None, custom_styles=None):
        if d["cod_componente"] == "ID-2":
            raise OSError("simulated render failure")
        return real_gen(d, out_path, formato, map_opts=map_opts, custom_styles=custom_styles)

    monkeypatch.setattr(ub, "generar_imagen_ubicacion", flaky_gen)

    out_dir = tmp_path / "out"
    result = ub.handle_generar_ubicaciones({
        "excelPath": str(xlsx),
        "outputDir": str(out_dir),
        "formato": "vertical",
        "consolidado": False,
    })

    assert result["generados"] == 1
    assert result["fallidos"] == 1
    pdfs = [p.name for p in out_dir.glob("*.pdf")]
    assert pdfs == ["ID-1.pdf"]


def test_map_cache_key_with_zoom_and_provider() -> None:
    key_default = ub._map_cache_key(-12.0, -77.0, "vertical", preview=True)
    key_zoom15 = ub._map_cache_key(-12.0, -77.0, "vertical", preview=True, map_opts={"zoom": 15})
    key_google = ub._map_cache_key(-12.0, -77.0, "vertical", preview=True, map_opts={"provider": "google"})
    key_maptiler_a = ub._map_cache_key(
        -12.0, -77.0, "vertical", preview=True,
        map_opts={"provider": "maptiler", "zoom": 18, "api_key": "KEY-A"},
    )
    key_maptiler_b = ub._map_cache_key(
        -12.0, -77.0, "vertical", preview=True,
        map_opts={"provider": "maptiler", "zoom": 18, "api_key": "KEY-B"},
    )

    assert key_default != key_zoom15
    assert key_default != key_google
    assert key_zoom15 != key_google
    assert key_maptiler_a != key_maptiler_b

    assert key_zoom15[-3:] == ("osm", 15, "")
    assert key_google[-3:-1] == ("google", 18)


def test_parse_combined_coord_value_rejects_invalid_text() -> None:
    assert ub._parse_combined_coord_value("invalid") == (None, None)
    assert ub._parse_combined_coord_value("abc, def") == (None, None)


def test_parse_combined_coord_value_parses_numeric_pair() -> None:
    lat, lon = ub._parse_combined_coord_value("-12.0464, -77.0428")
    assert lat == pytest.approx(-12.0464)
    assert lon == pytest.approx(-77.0428)


def test_parse_combined_coord_value_uses_latitude_longitude_order() -> None:
    lat, lon = ub._parse_combined_coord_value("-77.0428, -12.0464")
    assert lat == pytest.approx(-77.0428)
    assert lon == pytest.approx(-12.0464)


def test_parse_combined_coord_value_parses_google_maps_url() -> None:
    url = "https://maps.google.com/?q=-12.0464,-77.0428"
    lat, lon = ub._parse_combined_coord_value(url)
    assert lat == pytest.approx(-12.0464)
    assert lon == pytest.approx(-77.0428)


def test_handle_generar_ubicaciones_rejects_all_invalid_rows(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path,
) -> None:
    import openpyxl

    monkeypatch.setattr(ub, "fetch_static_map", lambda *a, **k: _png_bytes((256, 256)))

    xlsx = tmp_path / "invalid.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["coordenadas"])
    ws.append(["invalid"])
    ws.append(["also bad"])
    wb.save(xlsx)

    with pytest.raises(ValueError, match="coordenadas validas"):
        ub.handle_generar_ubicaciones({
            "excelPath": str(xlsx),
            "outputDir": str(tmp_path / "out"),
            "formato": "vertical",
            "consolidado": False,
        })


def test_handle_generar_ubicaciones_rejects_manual_invalid_coords(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="coordenadas validas"):
        ub.handle_generar_ubicaciones({
            "outputDir": str(tmp_path / "out"),
            "formato": "vertical",
            "manualData": {"lat": "abc", "lon": "def", "cod_componente": "X"},
        })


def test_handle_generar_ubicaciones_unique_pdf_for_duplicate_cod(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path,
) -> None:
    import openpyxl

    monkeypatch.setattr(ub, "fetch_static_map", lambda *a, **k: _png_bytes((256, 256)))

    xlsx = tmp_path / "dup.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["cod", "latitud", "longitud"])
    ws.append(["SAME", -12.0, -77.0])
    ws.append(["SAME", -13.0, -78.0])
    wb.save(xlsx)

    out_dir = tmp_path / "out"
    result = ub.handle_generar_ubicaciones({
        "excelPath": str(xlsx),
        "outputDir": str(out_dir),
        "formato": "vertical",
        "consolidado": False,
    })

    assert result["generados"] == 2
    pdfs = sorted(p.name for p in out_dir.glob("*.pdf"))
    assert pdfs == ["SAME.pdf", "SAME_2.pdf"]


@pytest.mark.parametrize("zoom", [-1, 23, 100_000_000, 1.5, "18", True])
def test_handle_preview_ubicacion_rejects_invalid_zoom(zoom: object) -> None:
    with pytest.raises(ValueError, match=r"(?i)zoom"):
        ub.handle_preview_ubicacion({
            "formato": "vertical",
            "zoom": zoom,
            "manualData": {"lat": "-12.0", "lon": "-77.0", "cod_componente": "ZOOM"},
        })


@pytest.mark.parametrize("zoom", [-1, 23, 100_000_000, 1.5, "18", True])
def test_handle_generar_ubicaciones_rejects_invalid_zoom(tmp_path: Path, zoom: object) -> None:
    with pytest.raises(ValueError, match=r"(?i)zoom"):
        ub.handle_generar_ubicaciones({
            "outputDir": str(tmp_path / "out"),
            "formato": "vertical",
            "zoom": zoom,
            "manualData": {"lat": "-12.0", "lon": "-77.0", "cod_componente": "ZOOM"},
        })


def test_handle_generar_ubicaciones_applies_custom_styles_and_map_opts(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path,
) -> None:
    """generar_ubicaciones debe pasar customStyles y map_opts al renderer."""
    captured: list[tuple] = []

    def fake_render(d, formato, map_opts=None, custom_styles=None):
        captured.append((formato, map_opts, custom_styles))
        img = Image.new("RGB", (100, 100), (10, 20, 30))
        return img

    monkeypatch.setattr(ub, "render_imagen_ubicacion", fake_render)

    out_dir = tmp_path / "out"
    styles = {"map": {"overlayAlpha": 50, "overlayColor": "#AABBCC"}}
    result = ub.handle_generar_ubicaciones({
        "outputDir": str(out_dir),
        "formato": "horizontal",
        "consolidado": True,
        "provider": "osm",
        "zoom": 16,
        "api_key": "",
        "customStyles": styles,
        "manualData": {
            "lat": "-12.0",
            "lon": "-77.0",
            "cod_componente": "GEN-1",
        },
    })

    assert result["consolidado"] is True
    assert (out_dir / "ubicaciones_consolidado.pdf").is_file()
    assert len(captured) == 1
    formato, map_opts, custom_styles = captured[0]
    assert formato == "horizontal"
    assert map_opts["provider"] == "osm"
    assert map_opts["zoom"] == 16
    assert custom_styles == styles


def _make_single_page_pdfs(tmp_path: Path, count: int = 1) -> list[str]:
    """Create temp single-page PDFs for merge tests."""
    paths = []
    for i in range(count):
        img = Image.new("RGB", (10, 10), (1, 2, 3))
        p = str(tmp_path / f"page_{i}.pdf")
        img.save(p, "PDF")
        img.close()
        paths.append(p)
    return paths


def test_merge_consolidated_pdfs_falls_back_when_default_locked(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path,
) -> None:
    """Si ubicaciones_consolidado.pdf está bloqueado, guardar en nombre alternativo."""
    page_paths = _make_single_page_pdfs(tmp_path, 1)
    original_replace = os.replace

    def fake_replace(src: str, dst: str) -> None:
        if dst.endswith("ubicaciones_consolidado.pdf"):
            raise PermissionError(13, "Permission denied", dst)
        original_replace(src, dst)

    monkeypatch.setattr(os, "replace", fake_replace)

    saved_path = ub._merge_consolidated_pdfs(page_paths, str(tmp_path))

    assert saved_path.endswith("ubicaciones_consolidado_2.pdf")
    assert (tmp_path / "ubicaciones_consolidado_2.pdf").is_file()


def test_merge_consolidated_pdfs_falls_back_on_windows_sharing_violation(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path,
) -> None:
    """WinError 32 (archivo en uso) también debe probar ubicaciones_consolidado_2.pdf."""
    page_paths = _make_single_page_pdfs(tmp_path, 1)
    original_replace = os.replace

    def fake_replace(src: str, dst: str) -> None:
        if dst.endswith("ubicaciones_consolidado.pdf"):
            err = OSError(0, "The process cannot access the file")
            err.winerror = 32  # type: ignore[attr-defined]
            raise err
        original_replace(src, dst)

    monkeypatch.setattr(os, "replace", fake_replace)

    saved_path = ub._merge_consolidated_pdfs(page_paths, str(tmp_path))

    assert saved_path.endswith("ubicaciones_consolidado_2.pdf")
    assert (tmp_path / "ubicaciones_consolidado_2.pdf").is_file()


def test_merge_consolidated_pdfs_raises_clear_error_when_all_paths_locked(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path,
) -> None:
    page_paths = _make_single_page_pdfs(tmp_path, 1)

    def always_denied(_src: str, _dst: str) -> None:
        raise PermissionError(13, "Permission denied", _dst)

    monkeypatch.setattr(os, "replace", always_denied)

    with pytest.raises(PermissionError, match="Cierra el archivo"):
        ub._merge_consolidated_pdfs(page_paths, str(tmp_path))


def test_merge_consolidated_pdfs_produces_multipage_pdf(tmp_path: Path) -> None:
    """Merge single-page temp PDFs into one multi-page PDF; clean up temps."""
    page_paths = _make_single_page_pdfs(tmp_path, 3)

    result = ub._merge_consolidated_pdfs(page_paths, str(tmp_path))

    assert os.path.exists(result)
    reader = PdfReader(result)
    assert len(reader.pages) == 3
    for p in page_paths:
        assert not os.path.exists(p), f"temp page file not cleaned up: {p}"


def test_consolidated_export_uses_managed_temp_directory(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    rendered_paths: list[Path] = []

    def fake_generate(_data: dict, output_path: str, _formato: str, *, map_opts=None, custom_styles=None) -> None:
        rendered_paths.append(Path(output_path))
        Image.new("RGB", (10, 10), (1, 2, 3)).save(output_path, "PDF")

    def fake_load_excel_data(_path: str):
        df = pd.DataFrame({
            "cod": ["1", "2", "3"],
            "latitud": [1.0, 2.0, 3.0],
            "longitud": [1.0, 2.0, 3.0],
        })
        return df, ("cod", None, None, None, "latitud", "longitud")

    monkeypatch.setattr(ub, "generar_imagen_ubicacion", fake_generate)
    monkeypatch.setattr(ub, "_load_excel_data", fake_load_excel_data)
    monkeypatch.setattr(ub.tempfile, "tempdir", str(tmp_path))

    result = ub.handle_generar_ubicaciones({
        "excelPath": str(tmp_path / "fake.xlsx"),
        "outputDir": str(tmp_path / "out"),
        "consolidado": True,
    })

    assert len(PdfReader(result["consolidatedPath"]).pages) == 3
    assert rendered_paths
    assert all(path.parent.name.startswith("antares-ubicaciones-") for path in rendered_paths)
    assert not list(tmp_path.glob("antares-ubicaciones-*"))


def test_consolidated_mode_closes_rendered_images(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    """Consolidated mode must close each PIL image after saving to a temp page
    PDF — holding all images in RAM OOMs on large batches."""
    closed = [0]
    held: list = []  # strong refs prevent GC → __del__ → close()

    def _track(img: Image.Image) -> None:
        held.append(img)
        orig_close = img.close
        orig_convert = img.convert

        def tracked_close() -> None:
            closed[0] += 1
            orig_close()

        def tracked_convert(mode: str, *a, **kw):  # type: ignore[no-untyped-def]
            converted = orig_convert(mode, *a, **kw)
            _track(converted)
            return converted

        img.close = tracked_close  # type: ignore[method-assign]
        img.convert = tracked_convert  # type: ignore[method-assign]

    def fake_render(datos: dict, formato: str, *, map_opts=None, custom_styles=None) -> Image.Image:
        img = Image.new("RGB", (10, 10), (255, 0, 0))
        _track(img)
        return img

    def fake_load_excel_data(path: str):
        df = pd.DataFrame({
            "cod": ["1", "2", "3"],
            "direccion": ["A", "B", "C"],
            "distrito": ["D1", "D2", "D3"],
            "latitud": [1.0, 2.0, 3.0],
            "longitud": [1.0, 2.0, 3.0],
        })
        return df, ("cod", "direccion", None, "distrito", "latitud", "longitud")

    monkeypatch.setattr(ub, "render_imagen_ubicacion", fake_render)
    monkeypatch.setattr(ub, "_load_excel_data", fake_load_excel_data)

    ub.handle_generar_ubicaciones({
        "excelPath": str(tmp_path / "fake.xlsx"),
        "outputDir": str(tmp_path),
        "consolidado": True,
    })

    assert closed[0] >= 3, f"expected ≥3 images closed (one per page), got {closed[0]}"


def test_consolidated_mode_failure_removes_temp_page_and_closes_images(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path,
) -> None:
    """Si rgb.save falla en modo consolidado, la página temporal debe
    eliminarse y las imágenes cerrarse — sin fugas en lotes con filas fallidas."""

    class Boom(Exception):
        pass

    closed = [0]

    def raising_save(_self, *_a, **_kw) -> None:  # type: ignore[no-untyped-def]
        raise Boom("save failed")

    def fake_render(datos: dict, formato: str, *, map_opts=None, custom_styles=None) -> Image.Image:
        img = Image.new("RGB", (10, 10), (255, 0, 0))
        orig_close = img.close

        def tracked_close() -> None:
            closed[0] += 1
            orig_close()

        img.close = tracked_close  # type: ignore[method-assign]
        return img

    real_mkstemp = tempfile.mkstemp

    def fake_mkstemp(**kw) -> tuple[int, str]:
        # Capture the real mkstemp before patching: tempfile.mkstemp is patched
        # at module level, so calling it by name here would recurse forever.
        kw.pop("dir", None)
        return real_mkstemp(dir=str(tmp_path), **kw)

    def fake_load_excel_data(path: str):
        df = pd.DataFrame({
            "cod": ["1", "2", "3"],
            "direccion": ["A", "B", "C"],
            "distrito": ["D1", "D2", "D3"],
            "latitud": [1.0, 2.0, 3.0],
            "longitud": [1.0, 2.0, 3.0],
        })
        return df, ("cod", "direccion", None, "distrito", "latitud", "longitud")

    monkeypatch.setattr(ub, "render_imagen_ubicacion", fake_render)
    monkeypatch.setattr(ub, "_load_excel_data", fake_load_excel_data)
    monkeypatch.setattr(ub.tempfile, "mkstemp", fake_mkstemp)
    monkeypatch.setattr(Image.Image, "save", raising_save)

    result = ub.handle_generar_ubicaciones({
        "excelPath": str(tmp_path / "fake.xlsx"),
        "outputDir": str(tmp_path),
        "consolidado": True,
    })

    assert result["fallidos"] == 3
    leftovers = list(tmp_path.glob("antares_page_*"))
    assert leftovers == [], f"temp page leaked after save failure: {leftovers}"
    assert closed[0] >= 3, f"expected ≥3 images closed, got {closed[0]}"


def test_preview_composed_cache_differs_by_formato() -> None:
    """Cambiar formato debe producir previews compuestas distintas."""
    manual = {
        "lat": "-11.968674",
        "lon": "-76.978299",
        "cod_componente": "FMT",
        "direccion": "Calle 1",
        "localidad": "Loc",
        "distrito": "Dist",
    }

    def fake_map(lat, lon, formato, preview=True, map_opts=None):
        w, h = ub._map_capture_size(formato, preview=preview)
        return _png_bytes((w, h), (80, 120, 160))

    with patch.object(ub, "_get_cached_map_screenshot", side_effect=fake_map):
        vertical = ub.handle_preview_ubicacion({"formato": "vertical", "manualData": manual})
        horizontal = ub.handle_preview_ubicacion({"formato": "horizontal", "manualData": manual})

    assert vertical["image"] != horizontal["image"]


@pytest.mark.parametrize(
    ("provider", "needle"),
    [
        ("mapbox", "access_token="),
        ("maptiler", "key="),
        ("stadia", "api_key="),
        ("geoapify", "apiKey="),
        ("thunderforest", "apikey="),
    ],
)
def test_xyz_providers_embed_api_key_in_tile_url(
    monkeypatch: pytest.MonkeyPatch, provider: str, needle: str,
) -> None:
    tile = _png_bytes((ub._OSM_TILE_SIZE, ub._OSM_TILE_SIZE), (60, 120, 160))
    seen: list[str] = []

    def fake_get(url: str, headers: dict[str, str], timeout: int = ub._HTTP_TIMEOUT) -> bytes:
        seen.append(url)
        return tile

    monkeypatch.setattr(ub, "_http_get", fake_get)
    ub.fetch_static_map(-12.046, -77.042, 400, 300, zoom=16, provider=provider, api_key="TEST-KEY-123")
    assert seen
    assert any(needle in url and "TEST-KEY-123" in url for url in seen)

