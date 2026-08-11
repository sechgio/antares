"""Regresión: spreadsheet_parse con staging .tmp, fechas y format_hint."""

from __future__ import annotations

import base64
import io
import json
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import pytest

from backend.handlers.spreadsheet import spreadsheet_parse


def _write_xlsx(path: Path, rows: list[list[Any]]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    for row in rows:
        ws.append(list(row))
    wb.save(path)
    wb.close()


def test_parse_xlsx_with_tmp_suffix_via_format_hint(tmp_path: Path) -> None:
    """Staging histórico guardaba datos.xlsx.tmp; format_hint debe bastar."""
    staged = tmp_path / "token_datos.xlsx.tmp"
    _write_xlsx(
        staged,
        [
            ["SGIO", "DISTRITO"],
            ["123", "ATE"],
            ["456", "LIMA"],
        ],
    )

    result = spreadsheet_parse({"path": str(staged), "format_hint": "xlsx"})

    assert result["workbookName"]
    assert len(result["sheets"]) == 1
    assert result["sheets"][0]["rows"][0] == ["SGIO", "DISTRITO"]
    assert result["sheets"][0]["rows"][1] == ["123", "ATE"]


def test_parse_xlsx_with_tmp_suffix_via_token_name(tmp_path: Path) -> None:
    """Sin format_hint, el nombre original del token define el formato."""
    staged = tmp_path / "antares-staged_upload.tmp"
    _write_xlsx(
        staged,
        [
            ["OT", "ZONA"],
            ["9", "Norte"],
        ],
    )

    result = spreadsheet_parse(
        {
            "_resolved_file_token_path": str(staged),
            "_resolved_file_token_name": "reporte.xlsx",
        },
    )

    assert result["workbookName"] == "reporte.xlsx"
    assert result["sheets"][0]["rows"][1] == ["9", "Norte"]


def test_parse_xlsx_datetime_cells_are_json_serializable(tmp_path: Path) -> None:
    """openpyxl puede devolver datetime; el resultado debe ser JSON-safe."""
    staged = tmp_path / "fechas.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["FECHA CORTE", "SGIO"])
    ws.append([datetime(2026, 3, 15, 8, 30, 0), "100"])
    ws.append([date(2026, 4, 1), "200"])
    wb.save(staged)
    wb.close()

    result = spreadsheet_parse({"path": str(staged), "format_hint": "xlsx"})

    json.dumps(result)  # no debe lanzar TypeError
    fecha_1 = result["sheets"][0]["rows"][1][0]
    fecha_2 = result["sheets"][0]["rows"][2][0]
    assert isinstance(fecha_1, str)
    assert isinstance(fecha_2, str)
    assert "2026" in fecha_1
    assert "2026" in fecha_2


def test_parse_xlsx_wrong_suffix_does_not_read_neighbor(tmp_path: Path) -> None:
    """Copia/renombre temporal no debe reutilizar un .xlsx vecino distinto."""
    neighbor = tmp_path / "upload.xlsx"
    _write_xlsx(neighbor, [["VECINO"], ["999"]])

    staged = tmp_path / "upload.bin"
    _write_xlsx(staged, [["ORIGEN"], ["1"]])

    result = spreadsheet_parse({"path": str(staged), "format_hint": "xlsx"})

    assert result["sheets"][0]["rows"][0] == ["ORIGEN"]
    assert result["sheets"][0]["rows"][1] == ["1"]


def test_parse_csv_via_hint_with_tmp_suffix(tmp_path: Path) -> None:
    staged = tmp_path / "datos.csv.tmp"
    staged.write_text("A,B\n1,2\n", encoding="utf-8-sig")

    result = spreadsheet_parse({"path": str(staged), "format_hint": "csv"})

    assert result["sheets"][0]["rows"] == [["A", "B"], ["1", "2"]]


def test_parse_spills_large_result_to_disk(tmp_path: Path, monkeypatch) -> None:
    """Large grids must not travel as inline JSON on the IPC pipe."""
    from backend.handlers import spreadsheet as ss

    monkeypatch.setattr(ss, "INLINE_RESULT_MAX_BYTES", 2_000)

    staged = tmp_path / "grande.xlsx"
    rows = [["C" + str(c) for c in range(20)] for _ in range(50)]
    rows[0] = [f"H{c}" for c in range(20)]
    _write_xlsx(staged, rows)

    result = spreadsheet_parse({"path": str(staged), "format_hint": "xlsx"})

    assert result["sheets"] == []
    assert "result_path" in result
    assert Path(result["result_path"]).is_file()
    spilled = json.loads(Path(result["result_path"]).read_text(encoding="utf-8"))
    assert len(spilled["sheets"][0]["rows"]) == 50
    assert result["sheet_meta"][0]["rowCount"] == 50
    assert result["workbookName"]


def test_get_rows_pages_spilled_cache(tmp_path: Path, monkeypatch) -> None:
    from backend.handlers import spreadsheet as ss
    from backend.handlers.spreadsheet import spreadsheet_get_rows

    monkeypatch.setattr(ss, "INLINE_RESULT_MAX_BYTES", 500)
    staged = tmp_path / "page.xlsx"
    rows = [[f"H{c}" for c in range(8)]] + [[str(r), "a", "b", "c", "d", "e", "f", "g"] for r in range(40)]
    _write_xlsx(staged, rows)

    parsed = spreadsheet_parse({"path": str(staged), "format_hint": "xlsx"})
    assert "result_path" in parsed, parsed
    page = spreadsheet_get_rows(
        {
            "result_path": parsed["result_path"],
            "sheet_index": 0,
            "offset": 0,
            "limit": 10,
        },
    )
    assert page["total"] == 41
    assert len(page["rows"]) == 10
    assert page["has_more"] is True
    assert page["rows"][0][0] == "H0"

    page2 = spreadsheet_get_rows(
        {
            "result_path": parsed["result_path"],
            "offset": 10,
            "limit": 10,
        },
    )
    assert len(page2["rows"]) == 10
    assert page2["rows"][0][0] == "9"


def test_b64_inline_uses_unique_temp_file() -> None:
    """b64 inline parse must use a unique temp file per call — a fixed name
    races between concurrent threads in the same process (scheduler pool)."""
    from backend.handlers.spreadsheet import _resolve_input_path

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["A", "B"])
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    b64 = base64.b64encode(buf.getvalue()).decode("ascii")

    p1, is_temp1 = _resolve_input_path({"xlsx_b64": b64})
    p2, is_temp2 = _resolve_input_path({"xlsx_b64": b64})

    try:
        assert is_temp1 and is_temp2, "b64 inline payloads must be flagged as temp"
        assert p1 != p2, "concurrent b64 parses must not share a fixed temp filename"
        assert p1.exists()
        assert p2.exists()
    finally:
        p1.unlink(missing_ok=True)
        p2.unlink(missing_ok=True)


def test_user_file_with_temp_prefix_not_deleted() -> None:
    """A real user file whose name starts with the temp prefix must NOT be
    unlinked by spreadsheet_parse — only files created for b64 payloads are
    (provenance, not name sniffing)."""
    from backend.handlers.spreadsheet import _INLINE_TEMP_PREFIX, spreadsheet_parse

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["A", "B"])
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    b64 = base64.b64encode(buf.getvalue()).decode("ascii")

    tmp_dir = tempfile.mkdtemp()
    try:
        user_file = Path(tmp_dir) / f"{_INLINE_TEMP_PREFIX}ventas.xlsx"
        user_file.write_bytes(base64.b64decode(b64))

        result = spreadsheet_parse({"path": str(user_file), "format_hint": "xlsx"})

        assert result["sheets"][0]["rows"][0] == ["A", "B"]
        assert user_file.exists(), "user file with temp-like name must survive parse"
    finally:
        import shutil

        shutil.rmtree(tmp_dir, ignore_errors=True)


def test_b64_inline_temp_file_removed_after_parse() -> None:
    """The unique temp file created for b64 inline parses must be deleted
    after parsing — unique names mean they would otherwise accumulate in
    %TEMP% for the app's lifetime."""
    from backend.handlers.spreadsheet import _INLINE_TEMP_PREFIX, spreadsheet_parse

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["A", "B"])
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    b64 = base64.b64encode(buf.getvalue()).decode("ascii")

    before = set(Path(tempfile.gettempdir()).glob(f"{_INLINE_TEMP_PREFIX}*"))
    result = spreadsheet_parse({"xlsx_b64": b64})
    after = set(Path(tempfile.gettempdir()).glob(f"{_INLINE_TEMP_PREFIX}*"))

    assert result["sheets"][0]["rows"][0] == ["A", "B"]
    assert after == before, f"b64 inline temp files leaked: {after - before}"


def test_b64_inline_parse_succeeds_even_if_temp_unlink_locked(monkeypatch) -> None:
    """A Windows lock on the temp file (PermissionError from unlink) must not
    fail an already-successful parse — cleanup is best-effort."""
    from backend.handlers import spreadsheet as ss

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["A", "B"])
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    b64 = base64.b64encode(buf.getvalue()).decode("ascii")

    def locked_unlink(_self, *_a, **_kw) -> None:
        raise PermissionError(13, "Permission denied", "antares_inline_x")

    monkeypatch.setattr(Path, "unlink", locked_unlink)

    result = ss.spreadsheet_parse({"xlsx_b64": b64})

    assert result["sheets"][0]["rows"][0] == ["A", "B"]


def test_b64_inline_write_failure_removes_temp_file(monkeypatch) -> None:
    """Si os.write falla (p.ej. disco lleno), el mkstemp no debe quedar
    huérfano — el cleanup ocurre dentro de _resolve_input_path."""
    from backend.handlers import spreadsheet as ss

    b64 = base64.b64encode(b"x" * 64).decode("ascii")

    def failing_write(_fd: int, _data: bytes) -> int:
        raise OSError(28, "No space left on device")

    monkeypatch.setattr(ss.os, "write", failing_write)

    before = set(Path(tempfile.gettempdir()).glob(f"{ss._INLINE_TEMP_PREFIX}*"))
    with pytest.raises(OSError, match="No space"):
        ss._resolve_input_path({"xlsx_b64": b64})
    after = set(Path(tempfile.gettempdir()).glob(f"{ss._INLINE_TEMP_PREFIX}*"))

    assert after == before, f"temp file leaked after write failure: {after - before}"


def test_b64_inline_rejects_oversized_payload(monkeypatch) -> None:
    """A b64 payload whose decoded size exceeds MAX_SPREADSHEET_BYTES must be
    rejected before decode to prevent OOM (same guard as the file-path branch)."""
    from backend.handlers import spreadsheet as ss

    monkeypatch.setattr(ss, "MAX_SPREADSHEET_BYTES", 100)
    oversized_b64 = base64.b64encode(b"x" * 200).decode("ascii")

    with pytest.raises(ValueError, match="excede"):
        ss._resolve_input_path({"xlsx_b64": oversized_b64})
